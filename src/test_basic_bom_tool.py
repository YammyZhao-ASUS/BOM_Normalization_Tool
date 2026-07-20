import tempfile
import unittest
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from basic_bom_tool import BasicBOMTool


class BasicBOMToolTests(unittest.TestCase):
    def test_generates_a_four_sheet_basic_report(self):
        bom = pd.DataFrame(
            [
                {
                    "Part Number": "C-A",
                    "Part Reference": "C1",
                    "Component_Name": "MLCC 100NF 16V 0402 X7R 10%",
                    "Vendor": "Murata",
                    "Quantity": 2,
                },
                {
                    "Part Number": "C-B",
                    "Part Reference": "C2",
                    "Component_Name": "MLCC 104 16V 0402 X7R 10%",
                    "Vendor": "TDK",
                    "Quantity": 1,
                },
                {
                    "Part Number": "R-A",
                    "Part Reference": "R1",
                    "Component_Name": "RES 10K 1/16W 0402 1% THICK FILM",
                },
                {
                    "Part Number": "R-B",
                    "Part Reference": "R2",
                    "Component_Name": "RES 12K 1/16W 0402 1% THICK FILM",
                },
            ]
        )
        tool = BasicBOMTool()
        reports = tool.analyze_dataframe(bom)

        self.assertEqual(len(reports["same_specification"]), 1)
        self.assertGreaterEqual(len(reports["near_values"]), 1)
        self.assertEqual(
            reports["summary"].set_index("Metric").loc["BOM Lines", "Value"],
            4,
        )

        with tempfile.TemporaryDirectory() as temporary_directory:
            output_file = Path(temporary_directory) / "basic_report.xlsx"
            tool.write_excel_report(reports, output_file)
            workbook = load_workbook(output_file, read_only=True)
            self.assertEqual(
                workbook.sheetnames,
                ["Summary", "Normalized BOM", "Same Specification", "Near Values"],
            )
            workbook.close()


if __name__ == "__main__":
    unittest.main()