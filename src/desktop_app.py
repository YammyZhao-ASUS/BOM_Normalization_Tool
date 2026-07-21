import os
import queue
import subprocess
import sys
import tempfile
import threading
import traceback
from datetime import datetime
from pathlib import Path
import tkinter as tk
from tkinter import filedialog, messagebox, ttk

try:
    from .bom_intelligence import BOMIntelligencePlatform
    from .sixty_bom_annotator import SixtyBOMAnnotator
except ImportError:
    from bom_intelligence import BOMIntelligencePlatform
    from sixty_bom_annotator import SixtyBOMAnnotator


APP_VERSION = "2.0.0"
MODE_AUTO = "Auto detect"
MODE_ENTERPRISE = "Enterprise report"
MODE_60BOM = "ASUS 60BOM annotation"


def resource_path(relative_path):
    application_root = Path(
        getattr(sys, "_MEIPASS", Path(__file__).resolve().parents[1])
    )
    return application_root / relative_path


def default_output_directory():
    if getattr(sys, "frozen", False):
        return Path.home() / "Documents" / "BOM Intelligence Reports"
    return Path(__file__).resolve().parents[1] / "output"


def run_self_test():
    """Exercise packaged resources, analysis, and Excel output without opening a GUI."""
    try:
        import pandas as pd
        from openpyxl import load_workbook

        rule_file = resource_path("config/default_rules.json")
        dataframe = pd.DataFrame(
            [
                {
                    "Part Number": "SELF-PN1",
                    "Part Reference": "C1",
                    "Component_Name": "MLCC 100NF 16V 0402 X7R 10%",
                    "Vendor": "Murata",
                    "Quantity": 2,
                },
                {
                    "Part Number": "SELF-PN2",
                    "Part Reference": "C2",
                    "Component_Name": "MLCC 104 16V 0402 X7R 10%",
                    "Vendor": "TDK",
                    "Quantity": 1,
                },
            ]
        )
        platform = BOMIntelligencePlatform(rule_file=rule_file)
        reports = platform.analyze_dataframe(dataframe)
        if len(reports["duplicate_pn"]) != 1:
            raise RuntimeError("Self-test duplicate specification analysis failed.")

        with tempfile.TemporaryDirectory() as temporary_directory:
            output_file = Path(temporary_directory) / "self_test_report.xlsx"
            platform.write_excel_report(reports, output_file)
            workbook = load_workbook(output_file, read_only=False, data_only=False)
            try:
                expected_sheets = [
                    "Overview",
                    "Merge Candidate",
                    "Capacitor Summary",
                    "Capacitor Detail",
                    "Resistor Summary",
                    "Resistor Detail",
                    "AVL Candidate",
                    "Risk Review",
                    "Nearby Value",
                    "Settings",
                ]
                if workbook.sheetnames != expected_sheets:
                    raise RuntimeError("Self-test workbook sheet count is invalid.")
                if workbook["Settings"].sheet_state != "hidden":
                    raise RuntimeError("Self-test settings sheet visibility is invalid.")
            finally:
                workbook.close()
        return 0
    except Exception:
        error_log = Path(tempfile.gettempdir()) / "BOM_Intelligence_self_test.log"
        error_log.write_text(traceback.format_exc(), encoding="utf-8")
        return 1


class BOMToolApp:
    """Tk desktop client for the BOM intelligence analysis services."""

    COLORS = {
        "navy": "#17324D",
        "teal": "#007F7B",
        "teal_dark": "#006B68",
        "paper": "#F4F7F9",
        "white": "#FFFFFF",
        "ink": "#17242F",
        "muted": "#607482",
        "line": "#D7E0E5",
    }

    SUMMARY_METRICS = (
        "Overall Score",
        "BOM Lines",
        "Duplicate PN Groups",
        "Near Resistance Pairs",
        "Near Capacitance Pairs",
        "Different Package Groups",
        "Different Voltage Groups",
        "Different Material Groups",
        "Unified AVL Ready",
        "Top Merge Candidates",
        "Cost Down Candidates",
        "High Risk Findings",
        "Average Data Quality",
    )

    def __init__(self, root):
        self.root = root
        self.root.title(f"BOM Intelligence Platform {APP_VERSION}")
        self.root.geometry("1040x740")
        self.root.minsize(900, 650)
        self.root.configure(bg=self.COLORS["paper"])
        self.root.protocol("WM_DELETE_WINDOW", self._on_close)

        source_default = resource_path("input/bom.xlsx")
        output_default = default_output_directory() / "BOM_Intelligence_Report.xlsx"
        rules_default = resource_path("config/default_rules.json")

        self.input_path = tk.StringVar(
            value=str(source_default) if source_default.is_file() else ""
        )
        self.output_path = tk.StringVar(value=str(output_default))
        self.rule_path = tk.StringVar(value=str(rules_default) if rules_default.is_file() else "")
        self.sheet_name = tk.StringVar()
        self.resistor_ratio = tk.StringVar(value="2.2")
        self.capacitor_ratio = tk.StringVar(value="2.2")
        self.analysis_mode = tk.StringVar(value=MODE_AUTO)
        self.status = tk.StringVar(value="Ready")
        self.status_detail = tk.StringVar(value="Select a BOM file and run analysis.")
        self.events = queue.Queue()
        self.running = False
        self.last_output = None

        self._configure_style()
        self._build_layout()
        self._bind_shortcuts()
        self.root.after(100, self._process_events)

    def _configure_style(self):
        style = ttk.Style(self.root)
        style.theme_use("clam")
        self.root.option_add("*Font", ("Segoe UI", 10))

        style.configure("App.TFrame", background=self.COLORS["paper"])
        style.configure("Header.TFrame", background=self.COLORS["navy"])
        style.configure(
            "Title.TLabel",
            background=self.COLORS["navy"],
            foreground=self.COLORS["white"],
            font=("Segoe UI Semibold", 20),
        )
        style.configure(
            "HeaderMeta.TLabel",
            background=self.COLORS["navy"],
            foreground="#C9D8E5",
            font=("Segoe UI", 9),
        )
        style.configure(
            "Section.TLabelframe",
            background=self.COLORS["white"],
            bordercolor=self.COLORS["line"],
            relief="solid",
        )
        style.configure(
            "Section.TLabelframe.Label",
            background=self.COLORS["white"],
            foreground=self.COLORS["navy"],
            font=("Segoe UI Semibold", 10),
        )
        style.configure("Panel.TFrame", background=self.COLORS["white"])
        style.configure(
            "Field.TLabel",
            background=self.COLORS["white"],
            foreground=self.COLORS["ink"],
            font=("Segoe UI Semibold", 9),
        )
        style.configure(
            "Hint.TLabel",
            background=self.COLORS["white"],
            foreground=self.COLORS["muted"],
            font=("Segoe UI", 9),
        )
        style.configure(
            "Accent.TButton",
            background=self.COLORS["teal"],
            foreground=self.COLORS["white"],
            bordercolor=self.COLORS["teal"],
            font=("Segoe UI Semibold", 10),
            padding=(18, 10),
        )
        style.map(
            "Accent.TButton",
            background=[("active", self.COLORS["teal_dark"]), ("disabled", "#9FB8B7")],
            bordercolor=[("active", self.COLORS["teal_dark"])],
        )
        style.configure("Action.TButton", padding=(12, 8))
        style.configure(
            "Status.TLabel",
            background=self.COLORS["paper"],
            foreground=self.COLORS["navy"],
            font=("Segoe UI Semibold", 10),
        )
        style.configure(
            "StatusDetail.TLabel",
            background=self.COLORS["paper"],
            foreground=self.COLORS["muted"],
            font=("Segoe UI", 9),
        )
        style.configure(
            "Treeview",
            background=self.COLORS["white"],
            fieldbackground=self.COLORS["white"],
            foreground=self.COLORS["ink"],
            rowheight=26,
            bordercolor=self.COLORS["line"],
        )
        style.configure(
            "Treeview.Heading",
            background="#E9F0F3",
            foreground=self.COLORS["navy"],
            font=("Segoe UI Semibold", 9),
            padding=(8, 7),
        )

    def _build_layout(self):
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(1, weight=1)

        header = ttk.Frame(self.root, style="Header.TFrame", padding=(28, 18))
        header.grid(row=0, column=0, sticky="ew")
        header.columnconfigure(0, weight=1)
        ttk.Label(header, text="BOM Intelligence Platform", style="Title.TLabel").grid(
            row=0,
            column=0,
            sticky="w",
        )
        ttk.Label(
            header,
            text="NORMALIZATION  |  AVL  |  COST  |  RISK",
            style="HeaderMeta.TLabel",
        ).grid(row=1, column=0, sticky="w", pady=(5, 0))
        ttk.Label(header, text=f"v{APP_VERSION}", style="HeaderMeta.TLabel").grid(
            row=0,
            column=1,
            rowspan=2,
            sticky="e",
        )

        content = ttk.Frame(self.root, style="App.TFrame", padding=(24, 20, 24, 10))
        content.grid(row=1, column=0, sticky="nsew")
        content.columnconfigure(0, weight=3)
        content.columnconfigure(1, weight=2)
        content.rowconfigure(0, weight=1)

        settings = ttk.LabelFrame(
            content,
            text="Analysis settings",
            style="Section.TLabelframe",
            padding=18,
        )
        settings.grid(row=0, column=0, sticky="nsew", padx=(0, 10))
        settings.columnconfigure(1, weight=1)

        self._add_path_row(settings, 0, "BOM file", self.input_path, self._choose_input)
        self._add_path_row(settings, 1, "Report file", self.output_path, self._choose_output)
        self._add_path_row(settings, 2, "Rule library", self.rule_path, self._choose_rules)

        ttk.Label(settings, text="Mode", style="Field.TLabel").grid(
            row=3,
            column=0,
            sticky="w",
            pady=(18, 6),
        )
        mode_box = ttk.Combobox(
            settings,
            textvariable=self.analysis_mode,
            values=(MODE_AUTO, MODE_ENTERPRISE, MODE_60BOM),
            state="readonly",
        )
        mode_box.grid(row=3, column=1, columnspan=2, sticky="ew", pady=(18, 6))

        ttk.Label(settings, text="Worksheet", style="Field.TLabel").grid(
            row=4,
            column=0,
            sticky="w",
            pady=6,
        )
        ttk.Entry(settings, textvariable=self.sheet_name).grid(
            row=4,
            column=1,
            columnspan=2,
            sticky="ew",
            pady=6,
        )
        ttk.Label(
            settings,
            text="Leave blank to detect the BOM worksheet and header automatically.",
            style="Hint.TLabel",
        ).grid(row=5, column=1, columnspan=2, sticky="w", pady=(0, 12))

        threshold_frame = ttk.Frame(settings, style="Panel.TFrame")
        threshold_frame.grid(row=6, column=0, columnspan=3, sticky="ew", pady=(7, 0))
        threshold_frame.columnconfigure(1, weight=1)
        threshold_frame.columnconfigure(3, weight=1)
        ttk.Label(threshold_frame, text="R near ratio", style="Field.TLabel").grid(
            row=0,
            column=0,
            sticky="w",
        )
        ttk.Spinbox(
            threshold_frame,
            from_=1.01,
            to=10.0,
            increment=0.05,
            textvariable=self.resistor_ratio,
            width=9,
        ).grid(row=0, column=1, sticky="w", padx=(10, 24))
        ttk.Label(threshold_frame, text="C near ratio", style="Field.TLabel").grid(
            row=0,
            column=2,
            sticky="w",
        )
        ttk.Spinbox(
            threshold_frame,
            from_=1.01,
            to=10.0,
            increment=0.05,
            textvariable=self.capacitor_ratio,
            width=9,
        ).grid(row=0, column=3, sticky="w", padx=(10, 0))

        self.run_button = ttk.Button(
            settings,
            text="Run analysis",
            style="Accent.TButton",
            command=self._start_analysis,
        )
        self.run_button.grid(row=7, column=0, columnspan=3, sticky="ew", pady=(26, 10))
        self.progress = ttk.Progressbar(settings, mode="indeterminate")
        self.progress.grid(row=8, column=0, columnspan=3, sticky="ew")

        results = ttk.LabelFrame(
            content,
            text="Latest analysis",
            style="Section.TLabelframe",
            padding=12,
        )
        results.grid(row=0, column=1, sticky="nsew", padx=(10, 0))
        results.columnconfigure(0, weight=1)
        results.rowconfigure(0, weight=1)

        self.summary_tree = ttk.Treeview(
            results,
            columns=("metric", "value"),
            show="headings",
            selectmode="none",
        )
        self.summary_tree.heading("metric", text="Metric")
        self.summary_tree.heading("value", text="Value")
        self.summary_tree.column("metric", width=190, minwidth=140, anchor="w")
        self.summary_tree.column("value", width=95, minwidth=70, anchor="e")
        self.summary_tree.grid(row=0, column=0, sticky="nsew")
        summary_scroll = ttk.Scrollbar(
            results,
            orient="vertical",
            command=self.summary_tree.yview,
        )
        summary_scroll.grid(row=0, column=1, sticky="ns")
        self.summary_tree.configure(yscrollcommand=summary_scroll.set)

        actions = ttk.Frame(results, style="Panel.TFrame")
        actions.grid(row=1, column=0, columnspan=2, sticky="ew", pady=(12, 0))
        actions.columnconfigure(0, weight=1)
        actions.columnconfigure(1, weight=1)
        self.open_report_button = ttk.Button(
            actions,
            text="Open report",
            style="Action.TButton",
            command=self._open_report,
            state="disabled",
        )
        self.open_report_button.grid(row=0, column=0, sticky="ew", padx=(0, 5))
        self.open_folder_button = ttk.Button(
            actions,
            text="Open folder",
            style="Action.TButton",
            command=self._open_output_folder,
            state="disabled",
        )
        self.open_folder_button.grid(row=0, column=1, sticky="ew", padx=(5, 0))

        activity = ttk.LabelFrame(
            self.root,
            text="Activity",
            style="Section.TLabelframe",
            padding=(14, 10),
        )
        activity.grid(row=2, column=0, sticky="ew", padx=24, pady=(0, 10))
        activity.columnconfigure(0, weight=1)
        self.activity_text = tk.Text(
            activity,
            height=4,
            borderwidth=0,
            background=self.COLORS["white"],
            foreground=self.COLORS["muted"],
            font=("Consolas", 9),
            wrap="word",
            state="disabled",
        )
        self.activity_text.grid(row=0, column=0, sticky="ew")

        status_bar = ttk.Frame(self.root, style="App.TFrame", padding=(24, 0, 24, 16))
        status_bar.grid(row=3, column=0, sticky="ew")
        status_bar.columnconfigure(1, weight=1)
        ttk.Label(status_bar, textvariable=self.status, style="Status.TLabel").grid(
            row=0,
            column=0,
            sticky="w",
        )
        ttk.Label(
            status_bar,
            textvariable=self.status_detail,
            style="StatusDetail.TLabel",
        ).grid(row=0, column=1, sticky="e")

    def _add_path_row(self, parent, row, label, variable, command):
        ttk.Label(parent, text=label, style="Field.TLabel").grid(
            row=row,
            column=0,
            sticky="w",
            pady=6,
        )
        ttk.Entry(parent, textvariable=variable).grid(
            row=row,
            column=1,
            sticky="ew",
            padx=(12, 8),
            pady=6,
        )
        ttk.Button(parent, text="Browse", command=command).grid(
            row=row,
            column=2,
            sticky="e",
            pady=6,
        )

    def _bind_shortcuts(self):
        self.root.bind("<Control-o>", lambda _event: self._choose_input())
        self.root.bind("<Control-r>", lambda _event: self._start_analysis())

    def _choose_input(self):
        current = Path(self.input_path.get()).expanduser()
        selected = filedialog.askopenfilename(
            title="Select BOM file",
            initialdir=str(current.parent if current.parent.is_dir() else Path.cwd()),
            filetypes=[
                ("BOM files", "*.xlsx *.xlsm *.xls *.csv *.txt"),
                ("Excel workbooks", "*.xlsx *.xlsm *.xls"),
                ("Delimited text", "*.csv *.txt"),
                ("All files", "*.*"),
            ],
        )
        if not selected:
            return

        self.input_path.set(selected)
        output_directory = Path(self.output_path.get()).expanduser().parent
        self.output_path.set(str(output_directory / f"{Path(selected).stem}_BOM_Report.xlsx"))

    def _choose_output(self):
        current = Path(self.output_path.get()).expanduser()
        selected = filedialog.asksaveasfilename(
            title="Save BOM report",
            initialdir=str(current.parent),
            initialfile=current.name,
            defaultextension=".xlsx",
            filetypes=[("Excel workbook", "*.xlsx")],
        )
        if selected:
            self.output_path.set(selected)

    def _choose_rules(self):
        current = Path(self.rule_path.get()).expanduser() if self.rule_path.get() else Path.cwd()
        selected = filedialog.askopenfilename(
            title="Select JSON rule library",
            initialdir=str(current.parent if current.parent.is_dir() else Path.cwd()),
            filetypes=[("JSON rule library", "*.json"), ("All files", "*.*")],
        )
        if selected:
            self.rule_path.set(selected)

    def _start_analysis(self):
        if self.running:
            return

        try:
            settings = self._validated_settings()
        except (ValueError, FileNotFoundError) as error:
            messagebox.showerror("BOM Intelligence Platform", str(error))
            return

        if settings["output_file"].exists() and not messagebox.askyesno(
            "Replace report",
            f"The report already exists. Replace it?\n\n{settings['output_file']}",
        ):
            return

        self.running = True
        self.run_button.state(["disabled"])
        self.open_report_button.state(["disabled"])
        self.open_folder_button.state(["disabled"])
        self.progress.start(12)
        self.status.set("Analyzing")
        self.status_detail.set("Parsing and normalizing BOM data...")
        self._append_activity(f"Started: {settings['input_file']}")

        worker = threading.Thread(
            target=self._run_analysis,
            kwargs=settings,
            daemon=True,
        )
        worker.start()

    def _validated_settings(self):
        input_file = Path(self.input_path.get()).expanduser().resolve()
        output_file = Path(self.output_path.get()).expanduser().resolve()
        rule_text = self.rule_path.get().strip()
        rule_file = Path(rule_text).expanduser().resolve() if rule_text else None

        if not input_file.is_file():
            raise FileNotFoundError("Choose an existing BOM file.")
        if output_file.suffix.casefold() != ".xlsx":
            raise ValueError("The report file must use the .xlsx extension.")
        if input_file == output_file:
            raise ValueError("The source BOM and report file must be different files.")
        if rule_file is not None and not rule_file.is_file():
            raise FileNotFoundError("Choose an existing JSON rule library or clear the field.")

        try:
            resistor_ratio = float(self.resistor_ratio.get())
            capacitor_ratio = float(self.capacitor_ratio.get())
        except ValueError as error:
            raise ValueError("Near-value ratios must be numeric.") from error
        if resistor_ratio <= 1 or capacitor_ratio <= 1:
            raise ValueError("Near-value ratios must be greater than 1.")

        mode = self.analysis_mode.get()
        if mode == MODE_60BOM and input_file.suffix.casefold() not in {".xlsx", ".xlsm"}:
            raise ValueError("ASUS 60BOM annotation requires an .xlsx or .xlsm workbook.")

        return {
            "input_file": input_file,
            "output_file": output_file,
            "rule_file": rule_file,
            "sheet_name": self.sheet_name.get().strip() or None,
            "resistor_ratio": resistor_ratio,
            "capacitor_ratio": capacitor_ratio,
            "mode": mode,
        }

    def _run_analysis(
        self,
        input_file,
        output_file,
        rule_file,
        sheet_name,
        resistor_ratio,
        capacitor_ratio,
        mode,
    ):
        try:
            use_60_bom = mode == MODE_60BOM
            if mode == MODE_AUTO and input_file.suffix.casefold() in {".xlsx", ".xlsm"}:
                use_60_bom = SixtyBOMAnnotator.supports_workbook(input_file)

            if use_60_bom:
                summary = SixtyBOMAnnotator(max(resistor_ratio, capacitor_ratio)).annotate(
                    input_file,
                    output_file,
                )
                metrics = {
                    "Electronic rows": summary["electronic_rows"],
                    "Near-value candidates": summary["near_value_candidates"],
                    "Second-source rows": summary["substitute_rows"],
                }
                result_mode = "ASUS 60BOM annotation"
            else:
                platform = BOMIntelligencePlatform(
                    near_value_ratio={"R": resistor_ratio, "C": capacitor_ratio},
                    rule_file=rule_file,
                )
                reports = platform.analyze_file(input_file, sheet_name=sheet_name)
                platform.write_excel_report(reports, output_file)
                summary = dict(zip(reports["summary"]["Metric"], reports["summary"]["Value"]))
                metrics = {
                    metric: summary[metric]
                    for metric in self.SUMMARY_METRICS
                    if metric in summary
                }
                result_mode = "Enterprise report"

            self.events.put(
                (
                    "success",
                    {
                        "output_file": output_file,
                        "metrics": metrics,
                        "mode": result_mode,
                    },
                )
            )
        except Exception as error:
            self.events.put(
                (
                    "error",
                    {
                        "message": str(error),
                        "detail": traceback.format_exc(),
                    },
                )
            )

    def _process_events(self):
        try:
            event, payload = self.events.get_nowait()
        except queue.Empty:
            self.root.after(100, self._process_events)
            return

        self.running = False
        self.progress.stop()
        self.run_button.state(["!disabled"])

        if event == "success":
            self.last_output = Path(payload["output_file"])
            self._populate_summary(payload["metrics"])
            self.status.set("Completed")
            self.status_detail.set(payload["mode"])
            self.open_report_button.state(["!disabled"])
            self.open_folder_button.state(["!disabled"])
            self._append_activity(f"Created: {self.last_output}")
            messagebox.showinfo(
                "BOM Intelligence Platform",
                f"Analysis completed.\n\n{self.last_output}",
            )
        else:
            self.status.set("Failed")
            self.status_detail.set(payload["message"])
            self._append_activity(payload["detail"].strip().splitlines()[-1])
            messagebox.showerror("BOM Intelligence Platform", payload["message"])

        self.root.after(100, self._process_events)

    def _populate_summary(self, metrics):
        for item in self.summary_tree.get_children():
            self.summary_tree.delete(item)
        for metric, value in metrics.items():
            formatted = f"{value:,.2f}" if isinstance(value, float) else f"{value:,}" if isinstance(value, int) else str(value)
            self.summary_tree.insert("", "end", values=(metric, formatted))

    def _append_activity(self, message):
        timestamp = datetime.now().strftime("%H:%M:%S")
        self.activity_text.configure(state="normal")
        self.activity_text.insert("end", f"[{timestamp}] {message}\n")
        self.activity_text.see("end")
        self.activity_text.configure(state="disabled")

    def _open_report(self):
        if self.last_output and self.last_output.is_file():
            self._open_path(self.last_output)

    def _open_output_folder(self):
        if self.last_output:
            self._open_path(self.last_output.parent)

    @staticmethod
    def _open_path(path):
        if sys.platform == "win32":
            os.startfile(str(path))
        elif sys.platform == "darwin":
            subprocess.Popen(["open", str(path)])
        else:
            subprocess.Popen(["xdg-open", str(path)])

    def _on_close(self):
        if self.running and not messagebox.askokcancel(
            "Exit",
            "Analysis is still running. Exit the application?",
        ):
            return
        self.root.destroy()


def main():
    root = tk.Tk()
    BOMToolApp(root)
    root.mainloop()


if __name__ == "__main__":
    if "--self-test" in sys.argv:
        raise SystemExit(run_self_test())
    main()