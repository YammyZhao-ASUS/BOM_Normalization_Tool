import tempfile
import unittest
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from bom_intelligence import BOMIntelligencePlatform
from rule_engine import RuleLibrary


class BOMIntelligencePlatformTests(unittest.TestCase):
    def setUp(self):
        self.bom = pd.DataFrame(
            [
                {
                    "Part Number": "PN1",
                    "Quantity": 2,
                    "Part Reference": "C1",
                    "Package_Type": "SMD",
                    "Component_Name": "MLCC 100NF 16V (0402) X7R",
                    "Description": "MURATA/GRM",
                    "Subsystem_ID": "PCH",
                },
                {
                    "Part Number": "PN2",
                    "Quantity": 1,
                    "Part Reference": "C2",
                    "Package_Type": "SMD",
                    "Component_Name": "MLCC 104 16V (0402) X7R",
                    "Description": "SAMSUNG/CL",
                    "Subsystem_ID": "PCH",
                },
                {
                    "Part Number": "PN3",
                    "Quantity": 1,
                    "Part Reference": "C3",
                    "Package_Type": "SMD",
                    "Component_Name": "MLCC 220NF 16V (0402) X7R",
                    "Description": "YAGEO/CC",
                    "Subsystem_ID": "PCH",
                },
                {
                    "Part Number": "PN4",
                    "Quantity": 1,
                    "Part Reference": "C4",
                    "Package_Type": "SMD",
                    "Component_Name": "MLCC 10UF 6.3V (0402) X5R",
                    "Description": "MURATA/GRM",
                    "Subsystem_ID": "P-CPU",
                },
                {
                    "Part Number": "PR1",
                    "Quantity": 1,
                    "Part Reference": "R1",
                    "Package_Type": "SMD",
                    "Component_Name": "RES 10K 1/16W 0402 1% THICK FILM",
                    "Description": "VISHAY/CRCW",
                    "Subsystem_ID": "PCH",
                },
                {
                    "Part Number": "PR2",
                    "Quantity": 1,
                    "Part Reference": "R2",
                    "Package_Type": "SMD",
                    "Component_Name": "RES 12K 1/16W 0402 1% THICK FILM",
                    "Description": "VISHAY/CRCW",
                    "Subsystem_ID": "PCH",
                },
            ]
        )
        self.platform = BOMIntelligencePlatform()

    def test_analysis_normalizes_and_protects_expected_rows(self):
        reports = self.platform.analyze_dataframe(self.bom)
        normalized_bom = reports["normalized_bom"]

        c2_value = normalized_bom.loc[
            normalized_bom["Reference"] == "C2", "Normalized_Value"
        ].iloc[0]
        self.assertEqual(c2_value, "100nF")
        self.assertEqual(len(reports["duplicate_pn"]), 1)
        self.assertGreaterEqual(len(reports["near_value"]), 2)
        self.assertEqual(len(reports["critical_parts"]), 1)
        self.assertGreaterEqual(len(reports["cost_down"]), 3)
        self.assertIn("Murata", reports["vendor_distribution"]["Vendor"].tolist())

    def test_report_contains_all_review_sheets(self):
        reports = self.platform.analyze_dataframe(self.bom)
        expected_sheets = [
            "Overview",
            "Merge Candidate",
            "Capacitor Summary",
            "Merge Workspace",
            "Resistor Summary",
            "Resistor Detail",
            "Resistor Nearby Value",
            "AVL Candidate",
            "Risk Review",
            "Settings",
        ]

        with tempfile.TemporaryDirectory() as temporary_directory:
            output_file = Path(temporary_directory) / "report.xlsx"
            self.platform.write_excel_report(reports, output_file)
            workbook = load_workbook(output_file, read_only=False)
            self.assertEqual(workbook.sheetnames, expected_sheets)
            self.assertEqual(workbook["Overview"]["A1"].value, "BOM Intelligence Overview")
            self.assertEqual(workbook["Settings"].sheet_state, "hidden")
            self.assertEqual(workbook["Merge Candidate"].freeze_panes, "A2")
            self.assertEqual(workbook["Capacitor Summary"].freeze_panes, "A2")
            self.assertEqual(workbook["Merge Workspace"].freeze_panes, "A2")
            self.assertEqual(workbook["Resistor Summary"].freeze_panes, "A2")
            self.assertEqual(workbook["Resistor Detail"].freeze_panes, "A2")
            self.assertEqual(workbook["Resistor Nearby Value"].freeze_panes, "A2")
            merge_headers = [cell.value for cell in workbook["Merge Candidate"][1]]
            summary_headers = [cell.value for cell in workbook["Capacitor Summary"][1]]
            detail_headers = [cell.value for cell in workbook["Merge Workspace"][1]]
            resistor_summary_headers = [cell.value for cell in workbook["Resistor Summary"][1]]
            resistor_detail_headers = [cell.value for cell in workbook["Resistor Detail"][1]]
            nearby_headers = [cell.value for cell in workbook["Resistor Nearby Value"][1]]
            self.assertEqual(
                merge_headers[:9],
                [
                    "Priority",
                    "Merge Difficulty",
                    "Difference",
                    "Spec",
                    "Merge PN",
                    "Merge Qty",
                    "Keep PN",
                    "Keep Qty",
                    "BOM Qty",
                ],
            )
            self.assertEqual(summary_headers[:9], ["Merge ID", "Review Item", "BOM Qty", "Keep Qty", "Merge Qty", "Priority", "Why Review", "BOM Action", "Detail"])
            self.assertEqual(detail_headers[:12], ["Merge ID", "Keep PN", "Merge PN", "Keep Qty", "Merge Qty", "Affected RefDes", "Estimated Modification", "Difference", "Package", "Voltage", "Material", "BOM Action"])
            self.assertNotIn("Vendor", detail_headers)
            self.assertEqual(resistor_summary_headers[:8], ["Value", "PN Count", "Total Qty", "Action / Target PN", "BOM Action", "Priority", "Reason (相似度分类)", "Group"])
            self.assertEqual(resistor_detail_headers[:11], ["Group", "Row Type", "Value", "Spec", "PN", "Qty", "Status", "Difference", "Why Listed", "BOM Action", "Merge Target"])
            self.assertEqual(nearby_headers[:8], ["Current Value", "Current BOM Qty", "Nearby Value", "Candidate Qty", "Difference", "Tolerance Band", "Family", "Candidate PNs"])
            self.assertNotIn("Recommendation", nearby_headers)
            self.assertIsNotNone(workbook["Capacitor Summary"]["I2"].hyperlink)
            self.assertIn("VLOOKUP", str(workbook["Capacitor Summary"]["H2"].value))
            self.assertTrue(workbook["Resistor Summary"].column_dimensions["H"].hidden)
            self.assertNotIn("RD Decision", resistor_summary_headers)
            self.assertFalse(self._has_data_validation(workbook["Capacitor Summary"], "H2:H1048576"))
            self.assertTrue(self._has_data_validation(workbook["Merge Workspace"], "L2:L1048576"))
            self.assertTrue(self._has_data_validation(workbook["Resistor Detail"], "J2:J1048576"))
            workbook.close()

    @staticmethod
    def _has_data_validation(worksheet, target_range):
        return any(
            target_range in str(validation.sqref)
            for validation in worksheet.data_validations.dataValidation
        )

    def test_capacitor_review_detail_expands_members_and_nested_differences(self):
        bom = pd.DataFrame(
            [
                {
                    "Part Number": "C-X5R-TARGET",
                    "Part Reference": "C10",
                    "Component_Name": "MLCC 100NF 16V 0402 X5R 10%",
                    "Vendor": "Murata",
                    "Quantity": 148,
                },
                {
                    "Part Number": "C-X6S-LOW",
                    "Part Reference": "C11",
                    "Component_Name": "MLCC 100NF 16V 0402 X6S 20%",
                    "Vendor": "TDK",
                    "Quantity": 60,
                },
            ]
        )
        reports = self.platform.analyze_dataframe(bom)

        with tempfile.TemporaryDirectory() as temporary_directory:
            output_file = Path(temporary_directory) / "capacitor_review.xlsx"
            self.platform.write_excel_report(reports, output_file)
            workbook = load_workbook(output_file, read_only=False)
            detail = workbook["Merge Workspace"]
            rows = [
                [detail.cell(row_index, column).value for column in range(1, 13)]
                for row_index in range(2, detail.max_row + 1)
                if detail.cell(row_index, 1).value
            ]
            keep_rows = [row for row in rows if not row[2]]
            merge_rows = [row for row in rows if row[2]]

            self.assertEqual(keep_rows[0][0], "M-001")
            self.assertEqual(keep_rows[0][1], "C-X5R-TARGET")
            self.assertEqual(keep_rows[0][3], 148)
            self.assertEqual(merge_rows[0][2], "C-X6S-LOW")
            self.assertEqual(merge_rows[0][4], 60)
            self.assertEqual(merge_rows[0][5], "C11")
            self.assertEqual(merge_rows[0][6], "Update 1 RefDes")
            self.assertIn("Vendor\nTDK → Murata", merge_rows[0][7])
            self.assertIn("X6S → X5R", merge_rows[0][10])
            self.assertIn(merge_rows[0][11], ("", None))
            self.assertEqual(float(keep_rows[0][3]) + float(merge_rows[0][4]), 208)
            workbook.close()

    def test_resistor_pages_are_qty_first_grouped_and_explain_why_listed(self):
        bom = pd.DataFrame(
            [
                {
                    "Part Number": "R100-LOW",
                    "Part Reference": "R10",
                    "Component_Name": "RES 100R 1/16W 0402 1% THICK FILM",
                    "Quantity": 2,
                },
                {
                    "Part Number": "R9760",
                    "Part Reference": "R11",
                    "Component_Name": "RES 9.76K 1/16W 0402 1% THICK FILM",
                    "Quantity": 4,
                },
                {
                    "Part Number": "R9900",
                    "Part Reference": "R15",
                    "Component_Name": "RES 9.9K 1/16W 0402 1% THICK FILM",
                    "Quantity": 10,
                },
                {
                    "Part Number": "R10K-TARGET",
                    "Part Reference": "R12",
                    "Component_Name": "RES 10K 1/16W 0402 1% THICK FILM",
                    "Vendor": "Yageo",
                    "Quantity": 98,
                },
                {
                    "Part Number": "R10K-LOW",
                    "Part Reference": "R13",
                    "Component_Name": "RES 10K 1/16W 0402 1% THICK FILM",
                    "Vendor": "Vishay",
                    "Quantity": 3,
                },
                {
                    "Part Number": "R10K-TOL",
                    "Part Reference": "R14",
                    "Component_Name": "RES 10K 1/16W 0402 5% THICK FILM",
                    "Quantity": 2,
                },
                {
                    "Part Number": "R10100",
                    "Part Reference": "R16",
                    "Component_Name": "RES 10.1K 1/16W 0402 1% THICK FILM",
                    "Quantity": 30,
                },
                {
                    "Part Number": "R10200",
                    "Part Reference": "R17",
                    "Component_Name": "RES 10.2K 1/16W 0402 1% THICK FILM",
                    "Quantity": 20,
                },
            ]
        )

        reports = self.platform.analyze_dataframe(bom)

        with tempfile.TemporaryDirectory() as temporary_directory:
            output_file = Path(temporary_directory) / "resistor_report.xlsx"
            self.platform.write_excel_report(reports, output_file)
            workbook = load_workbook(output_file, read_only=False)
            summary = workbook["Resistor Summary"]
            detail = workbook["Resistor Detail"]
            nearby = workbook["Resistor Nearby Value"]
            summary_rows = {
                summary.cell(row_index, 1).value: [summary.cell(row_index, column).value for column in range(1, 9)]
                for row_index in range(2, summary.max_row + 1)
                if summary.cell(row_index, 1).value
            }

            self.assertNotIn("100Ohm", summary_rows)
            self.assertIsNotNone(summary["A2"].hyperlink)
            self.assertEqual(summary_rows["10kOhm"][1], 3)
            self.assertEqual(summary_rows["10kOhm"][3], "Review Required")
            self.assertIn("VLOOKUP", summary_rows["10kOhm"][4])
            self.assertEqual(summary_rows["10kOhm"][5], "★★★")
            self.assertEqual(summary_rows["10kOhm"][6], "🟡 Tolerance 1%⇄5%")
            self.assertEqual(summary_rows["9.76kOhm"][3], "Review Required")
            self.assertIn("🟠 Near Value 9.76kΩ ➔ 9.9kΩ", summary_rows["9.76kOhm"][6])
            self.assertIn("Target Value : 9.9kΩ", summary_rows["9.76kOhm"][6])
            self.assertIn("Package : 0402", summary_rows["9.76kOhm"][6])
            self.assertIn("Qty : 10 pcs", summary_rows["9.76kOhm"][6])
            self.assertIn("Difference : -1.41%", summary_rows["9.76kOhm"][6])

            ten_k_pn_rows = [
                [detail.cell(row_index, column).value for column in range(1, 10)]
                for row_index in range(2, detail.max_row + 1)
                if detail.cell(row_index, 3).value == "10kOhm" and detail.cell(row_index, 2).value == "PN"
            ]
            self.assertEqual([row[5] for row in ten_k_pn_rows], [2, 3, 98])
            self.assertIn("Target", [row[6] for row in ten_k_pn_rows])
            ten_k_header = next(
                [detail.cell(row_index, column).value for column in range(1, 10)]
                for row_index in range(2, detail.max_row + 1)
                if detail.cell(row_index, 3).value == "10kOhm" and detail.cell(row_index, 2).value == "Value Header"
            )
            self.assertIn("10.1kΩ", ten_k_header[7])
            self.assertIn("10.2kΩ", ten_k_header[7])
            self.assertIn("9.9kΩ", ten_k_header[7])
            self.assertNotIn("9.76kΩ", ten_k_header[7])
            nine_76_candidate_rows = [
                [detail.cell(row_index, column).value for column in range(1, 10)]
                for row_index in range(2, detail.max_row + 1)
                if detail.cell(row_index, 3).value == "9.9kΩ" and detail.cell(row_index, 2).value in {"Candidate", "Candidate PN"}
            ]
            self.assertEqual(nine_76_candidate_rows[0][1], "Candidate")
            self.assertIn("Total Qty : 10 pcs", nine_76_candidate_rows[0][3])
            self.assertEqual(nine_76_candidate_rows[0][6], "Nearby Candidate")
            self.assertEqual(nine_76_candidate_rows[1][1], "Candidate PN")
            self.assertEqual(nine_76_candidate_rows[1][4], "R9900")
            self.assertEqual(nine_76_candidate_rows[1][5], 10)
            self.assertEqual(nine_76_candidate_rows[1][6], "Nearby Candidate")
            self.assertEqual([cell.value for cell in nearby[1]], ["Current Value", "Current BOM Qty", "Nearby Value", "Candidate Qty", "Difference", "Tolerance Band", "Family", "Candidate PNs"])
            nearby_rows = [
                [nearby.cell(row_index, column).value for column in range(1, 9)]
                for row_index in range(2, nearby.max_row + 1)
            ]
            self.assertTrue(any(row[0] == "10kOhm" and row[2] == "10.2kOhm" and row[3] == 20 for row in nearby_rows))
            self.assertTrue(all(abs(float(str(row[4]).replace("%", "").replace("'", ""))) <= 10 for row in nearby_rows if row[4]))
            self.assertFalse(any(row[0] == "1.21kOhm" and row[2] == "2.2kOhm" for row in nearby_rows))
            self.assertNotIn("Recommendation", [cell.value for cell in nearby[1]])
            workbook.close()

    def test_merge_candidates_rank_low_quantity_pns_to_target_pn(self):
        bom = pd.DataFrame(
            [
                {
                    "Part Number": "R-TARGET",
                    "Part Reference": "R102",
                    "Component_Name": "RES 10K 1/16W 0402 1% THICK FILM",
                    "Vendor": "Yageo",
                    "Quantity": 128,
                },
                {
                    "Part Number": "R-LOW-1",
                    "Part Reference": "R153",
                    "Component_Name": "RES 10K 1/16W 0402 1% THICK FILM",
                    "Vendor": "Yageo",
                    "Quantity": 2,
                },
                {
                    "Part Number": "R-LOW-2",
                    "Part Reference": "R608",
                    "Component_Name": "RES 10K 1/16W 0402 1% THICK FILM",
                    "Vendor": "Vishay",
                    "Quantity": 5,
                },
                {
                    "Part Number": "R-LATER",
                    "Part Reference": "R777",
                    "Component_Name": "RES 10K 1/16W 0402 1% THICK FILM",
                    "Vendor": "Yageo",
                    "Quantity": 17,
                },
            ]
        )

        merge_candidates = self.platform.analyze_dataframe(bom)["merge_candidates"]

        self.assertEqual(merge_candidates["Current_PN"].tolist(), ["R-LOW-1", "R-LOW-2"])
        self.assertTrue((merge_candidates["Target_PN"] == "R-TARGET").all())
        self.assertEqual(merge_candidates.iloc[0]["Priority"], "Priority 1")
        self.assertEqual(merge_candidates.iloc[0]["Priority_Stars"], "★★★★★")
        self.assertEqual(merge_candidates.iloc[0]["Quantity_Score"], 100)
        self.assertEqual(merge_candidates.iloc[1]["Spec_Similarity"], 95)

    def test_merge_candidates_ignore_second_source_rows(self):
        bom = pd.DataFrame(
            [
                {
                    "Part Number": "R-TARGET",
                    "Part Reference": "R1",
                    "Item Description": "RES 10K 1/16W 0402 1% THICK FILM",
                    "Comp Type": "",
                    "Quantity": 100,
                },
                {
                    "Part Number": "R-SECOND-SOURCE",
                    "Part Reference": "R1-S",
                    "Item Description": "RES 10K 1/16W 0402 1% THICK FILM",
                    "Comp Type": "S",
                    "Quantity": 1,
                },
                {
                    "Part Number": "R-MAIN-LOW",
                    "Part Reference": "R2",
                    "Item Description": "RES 10K 1/16W 0402 1% THICK FILM",
                    "Comp Type": "",
                    "Quantity": 2,
                },
            ]
        )

        reports = self.platform.analyze_dataframe(bom)
        merge_candidates = reports["merge_candidates"]

        self.assertEqual(merge_candidates["Current_PN"].tolist(), ["R-MAIN-LOW"])
        self.assertNotIn("R-SECOND-SOURCE", merge_candidates["Current_PN"].tolist())
        normalized = reports["normalized_bom"].set_index("Part_Number")
        self.assertTrue(normalized.loc["R-SECOND-SOURCE", "Is_Second_Source"])

    def test_project_name_uses_first_item_description_prefix(self):
        bom = pd.DataFrame(
            [
                {
                    "Part Number": "ASM",
                    "Part Reference": "",
                    "Item Description": "P500MV MAIN BD MODULE//MECHANICAL",
                    "Quantity": 1,
                }
            ]
        )

        reports = self.platform.analyze_dataframe(bom)
        metadata = reports["report_metadata"].set_index("Property")["Value"]

        self.assertEqual(metadata["Project Name"], "P500MV")

    def test_finds_variants_avl_cost_and_risk_opportunities(self):
        rows = [
            ("C1", "PN1", "MLCC 100NF 16V 0402 X7R 10%", "Murata", 10, 0.05, "Active"),
            ("C2", "PN2", "MLCC 104 16V 0402 X7R 10%", "Samsung", 20, 0.02, "Active"),
            ("C3", "PN3", "MLCC 100NF 25V 0402 X7R 10%", "Murata", 1, None, "Active"),
            ("C4", "PN4", "MLCC 100NF 16V 0402 X5R 10%", "TDK", 1, None, "Active"),
            ("C5", "PN5", "MLCC 100NF 16V 0603 X7R 10%", "Yageo", 1, None, "Active"),
            ("R1", "PR1", "RES 10K OHM 1/16W 0402 1% THICK FILM", "Yageo", 1, None, "EOL"),
            ("R2", "PR2", "RES 12K OHM 1/16W 0402 1% THICK FILM", "Yageo", 1, None, "Active"),
        ]
        dataframe = pd.DataFrame(
            [
                {
                    "Part Reference": reference,
                    "Part Number": part_number,
                    "Component_Name": description,
                    "Vendor": vendor,
                    "Quantity": quantity,
                    "Unit Price": unit_price,
                    "Lifecycle": lifecycle,
                }
                for reference, part_number, description, vendor, quantity, unit_price, lifecycle in rows
            ]
        )

        reports = self.platform.analyze_dataframe(dataframe)

        self.assertEqual(len(reports["duplicate_pn"]), 1)
        self.assertEqual(len(reports["different_package"]), 1)
        self.assertEqual(len(reports["different_voltage"]), 1)
        self.assertEqual(len(reports["different_material"]), 1)
        self.assertEqual(len(reports["near_resistance"]), 1)
        self.assertIn("Ready for unified AVL", set(reports["avl_candidates"]["AVL_Readiness"]))
        self.assertIn("LIFE-001", set(reports["risk_components"]["Rule_ID"]))
        self.assertAlmostEqual(
            reports["duplicate_pn"].iloc[0]["Estimated_BOM_Savings"],
            0.3,
        )

    def test_report_escapes_formula_like_source_text(self):
        bom = self.bom.copy()
        bom.loc[0, "Part Number"] = "\t=1+1"
        bom["=HYPERLINK(\"bad\")"] = "source value"
        reports = self.platform.analyze_dataframe(bom)

        with tempfile.TemporaryDirectory() as temporary_directory:
            output_file = Path(temporary_directory) / "safe_report.xlsx"
            self.platform.write_excel_report(reports, output_file)
            workbook = load_workbook(output_file, data_only=False)
            formula_like_cells = [
                cell
                for worksheet in workbook.worksheets
                for row in worksheet.iter_rows()
                for cell in row
                if isinstance(cell.value, str) and "=1+1" in cell.value
            ]
            self.assertTrue(formula_like_cells)
            self.assertTrue(
                all(
                    str(cell.value).startswith("'")
                    for cell in formula_like_cells
                    if str(cell.value).lstrip("\t\r\n\ufeff").startswith("=1+1")
                )
            )
            workbook.close()

    def test_description_only_signal_context_is_protected(self):
        bom = pd.DataFrame(
            [
                {
                    "Part Number": "PR1",
                    "Part Reference": "R1",
                    "Component_Name": "RES 22R 1/16W 0402 1% THICK FILM",
                    "Description": "USB TX termination",
                }
            ]
        )

        normalized = self.platform.analyze_dataframe(bom)["normalized_bom"].iloc[0]

        self.assertTrue(normalized["Is_Critical"])
        self.assertEqual(normalized["Critical_Reason"], "High-speed or RF signal context")

    def test_avl_preferred_part_is_approved_and_vendor_identity_is_casefolded(self):
        rules = RuleLibrary().to_dict()
        rules["avl"]["approved_vendors"] = ["Murata", "TDK"]
        platform = BOMIntelligencePlatform(rule_library=RuleLibrary(rules))
        bom = pd.DataFrame(
            [
                {
                    "Part Number": "PN-UNAPPROVED",
                    "Part Reference": "C1",
                    "Component_Name": "MLCC 100NF 16V 0402 X7R 10%",
                    "Vendor": "CheapCo",
                    "Unit Price": 0.01,
                },
                {
                    "Part Number": "PN-MURATA",
                    "Part Reference": "C2",
                    "Component_Name": "MLCC 100NF 16V 0402 X7R 10%",
                    "Vendor": "MURATA",
                    "Unit Price": 0.05,
                },
                {
                    "Part Number": "PN-TDK",
                    "Part Reference": "C3",
                    "Component_Name": "MLCC 100NF 16V 0402 X7R 10%",
                    "Vendor": "tdk",
                    "Unit Price": 0.06,
                },
            ]
        )

        avl = platform.analyze_dataframe(bom)["avl_candidates"].iloc[0]

        self.assertEqual(avl["Preferred_PN"], "PN-MURATA")
        self.assertEqual(avl["Vendor_Count"], 3)
        self.assertEqual(avl["AVL_Readiness"], "Ready for unified AVL")

    def test_generic_material_does_not_hide_capacitor_dielectric_difference(self):
        bom = pd.DataFrame(
            [
                {
                    "Part Number": "PN-X5R",
                    "Part Reference": "C1",
                    "Component_Name": "MLCC 100NF 16V 0402 X5R 10%",
                    "Material": "Ceramic",
                    "Vendor": "Murata",
                },
                {
                    "Part Number": "PN-X7R",
                    "Part Reference": "C2",
                    "Component_Name": "MLCC 100NF 16V 0402 X7R 10%",
                    "Material": "Ceramic",
                    "Vendor": "TDK",
                },
            ]
        )

        reports = self.platform.analyze_dataframe(bom)

        self.assertTrue(reports["duplicate_pn"].empty)
        self.assertEqual(len(reports["different_material"]), 1)
        self.assertEqual(
            reports["different_material"].iloc[0]["Attribute_Values"],
            "X5R, X7R",
        )

    def test_incomplete_specifications_are_not_comparison_candidates(self):
        bom = pd.DataFrame(
            [
                {
                    "Part Number": "PR1",
                    "Part Reference": "R1",
                    "Component_Name": "RES 10K 0402",
                    "Vendor": "Yageo",
                },
                {
                    "Part Number": "PR2",
                    "Part Reference": "R2",
                    "Component_Name": "RES 12K 0402",
                    "Vendor": "Yageo",
                },
            ]
        )

        reports = self.platform.analyze_dataframe(bom)

        self.assertFalse(reports["normalized_bom"]["Comparison_Eligible"].any())
        self.assertTrue(reports["duplicate_pn"].empty)
        self.assertTrue(reports["near_resistance"].empty)
        self.assertTrue(reports["avl_candidates"].empty)
        self.assertIn("DATA-003", set(reports["ai_rule_findings"]["Rule_ID"]))

    def test_quantity_and_embedded_voltage_are_normalized_safely(self):
        bom = pd.DataFrame(
            [
                {
                    "Part Number": "PC1",
                    "Part Reference": "C1",
                    "Component_Name": "MLCC 10UF 6V3 0402 X5R 10%",
                    "Quantity": "1,000",
                    "Vendor": "Murata",
                },
                {
                    "Part Number": "PC2",
                    "Part Reference": "C2",
                    "Component_Name": "MLCC 10UF 6V3 0402 X5R 10%",
                    "Quantity": "inf",
                    "Vendor": "TDK",
                },
            ]
        )

        reports = self.platform.analyze_dataframe(bom)
        normalized = reports["normalized_bom"].set_index("Part_Number")

        self.assertEqual(normalized.loc["PC1", "Voltage"], "6.3V")
        self.assertEqual(normalized.loc["PC1", "Quantity_Normalized"], 1000)
        self.assertTrue(normalized.loc["PC1", "Quantity_Valid"])
        self.assertEqual(normalized.loc["PC2", "Quantity_Normalized"], 1)
        self.assertFalse(normalized.loc["PC2", "Quantity_Valid"])
        self.assertIn("DATA-004", set(reports["ai_rule_findings"]["Rule_ID"]))


if __name__ == "__main__":
    unittest.main()