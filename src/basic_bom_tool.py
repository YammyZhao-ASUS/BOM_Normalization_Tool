from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill

try:
    from .bom_intelligence import BOMIntelligencePlatform
except ImportError:
    from bom_intelligence import BOMIntelligencePlatform


class BasicBOMTool:
    """Small, working BOM normalization workflow for first-time users."""

    SHEETS = (
        ("Summary", "summary"),
        ("Normalized BOM", "normalized_bom"),
        ("Same Specification", "same_specification"),
        ("Near Values", "near_values"),
    )

    NORMALIZED_COLUMNS = (
        "Part_Number",
        "Reference",
        "Component_Type",
        "Original_Value",
        "Normalized_Value",
        "Voltage",
        "Dielectric",
        "Material",
        "Tolerance",
        "Power_Rating",
        "Package_Identity",
        "Vendor",
        "Quantity_Normalized",
        "Review_Status",
    )

    DUPLICATE_COLUMNS = (
        "Group_ID",
        "Component_Type",
        "Normalized_Value",
        "Voltage",
        "Dielectric",
        "Material",
        "Size",
        "Part_Number_Count",
        "Part_Numbers",
        "Vendors",
        "Total_Quantity",
        "References",
        "Recommendation",
    )

    NEAR_VALUE_COLUMNS = (
        "Pair_ID",
        "Component_Type",
        "Value_A",
        "Value_B",
        "Ratio",
        "Difference_Percent",
        "Voltage",
        "Dielectric",
        "Material",
        "Size",
        "References_A",
        "References_B",
        "Recommendation",
    )

    def __init__(self, near_value_ratio=2.2):
        self.platform = BOMIntelligencePlatform(near_value_ratio=near_value_ratio)

    def analyze_file(self, input_file):
        return self._to_basic_reports(self.platform.analyze_file(input_file))

    def analyze_dataframe(self, dataframe):
        return self._to_basic_reports(self.platform.analyze_dataframe(dataframe))

    def write_excel_report(self, reports, output_file):
        output_path = Path(output_file).expanduser().resolve()
        output_path.parent.mkdir(parents=True, exist_ok=True)

        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            for sheet_name, report_key in self.SHEETS:
                dataframe = self._sanitize_dataframe(reports[report_key])
                dataframe.to_excel(writer, sheet_name=sheet_name, index=False)
                self._format_sheet(writer.sheets[sheet_name], dataframe)

        return str(output_path)

    def _to_basic_reports(self, reports):
        normalized = self._select_columns(
            reports["normalized_bom"],
            self.NORMALIZED_COLUMNS,
        )
        same_specification = self._select_columns(
            reports["duplicate_pn"],
            self.DUPLICATE_COLUMNS,
        )
        near_values = self._select_columns(
            reports["near_value"],
            self.NEAR_VALUE_COLUMNS,
        )

        summary = pd.DataFrame(
            [
                ("BOM Lines", len(normalized)),
                ("Same Specification Groups", len(same_specification)),
                ("Near Value Pairs", len(near_values)),
            ],
            columns=["Metric", "Value"],
        )

        return {
            "summary": summary,
            "normalized_bom": normalized,
            "same_specification": same_specification,
            "near_values": near_values,
        }

    @staticmethod
    def _select_columns(dataframe, columns):
        available_columns = [column for column in columns if column in dataframe.columns]
        return dataframe.loc[:, available_columns].copy()

    def _sanitize_dataframe(self, dataframe):
        safe_dataframe = dataframe.copy()
        safe_dataframe.columns = self._safe_column_labels(safe_dataframe.columns)
        for column in safe_dataframe.columns:
            safe_dataframe[column] = safe_dataframe[column].map(self._safe_excel_value)
        return safe_dataframe

    @staticmethod
    def _safe_column_labels(columns):
        labels = []
        counts = {}
        for index, column in enumerate(columns, start=1):
            label = str(BasicBOMTool._safe_excel_value(str(column))).strip()
            if not label:
                label = f"Column_{index}"
            counts[label] = counts.get(label, 0) + 1
            labels.append(label if counts[label] == 1 else f"{label}_{counts[label]}")
        return labels

    @staticmethod
    def _safe_excel_value(value):
        if not isinstance(value, str):
            return value
        if value.lstrip("\t\r\n\ufeff").startswith(("=", "+", "-", "@")):
            return f"'{value}"
        return value

    @staticmethod
    def _format_sheet(worksheet, dataframe):
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions
        worksheet.row_dimensions[1].height = 24

        header_fill = PatternFill("solid", fgColor="1F4E78")
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for column_index, column_name in enumerate(dataframe.columns, start=1):
            values = [str(column_name)]
            values.extend(str(value) for value in dataframe[column_name].head(200).fillna(""))
            width = min(max(max(len(value) for value in values) + 2, 12), 42)
            column_letter = worksheet.cell(1, column_index).column_letter
            worksheet.column_dimensions[column_letter].width = width

        for row in worksheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)