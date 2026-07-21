import os
import re
import uuid
from pathlib import Path

import pandas as pd
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import ColorScaleRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


class ExcelReportWriter:
    """Create an RD-focused Excel workbook for daily PN optimization work."""

    SHEETS = (
        ("Merge Candidate", "merge_candidates"),
        ("Capacitor Summary", "specification_summary"),
        ("Merge Workspace", "specification_detail"),
        ("Resistor Summary", "resistor_summary"),
        ("Resistor Detail", "resistor_detail"),
        ("AVL Candidate", "avl_candidate"),
        ("Risk Review", "risk_review"),
        ("Resistor Nearby Value", "nearby_value"),
        ("Settings", "settings"),
    )

    COLORS = {
        "navy": "17324D",
        "teal": "007F7B",
        "cyan": "DDF4F2",
        "ink": "17242F",
        "muted": "607482",
        "line": "D7E0E5",
        "paper": "F5F8FA",
        "white": "FFFFFF",
        "critical": "F8D7DA",
        "high": "FCE5CD",
        "medium": "FFF2CC",
        "low": "D9EAD3",
    }

    def write(self, reports, output_file):
        output_path = Path(output_file).expanduser().resolve()
        output_path.parent.mkdir(parents=True, exist_ok=True)
        temporary_path = output_path.with_name(
            f".{output_path.stem}.{uuid.uuid4().hex}.tmp{output_path.suffix or '.xlsx'}"
        )

        try:
            with pd.ExcelWriter(temporary_path, engine="openpyxl") as writer:
                workbook = writer.book
                overview = workbook.create_sheet("Overview", 0)
                report_views = self._build_report_views(reports)

                for sheet_name, report_key in self.SHEETS:
                    dataframe = report_views.get(report_key, pd.DataFrame())
                    safe_dataframe = self._sanitize_dataframe(dataframe)
                    safe_dataframe.to_excel(writer, sheet_name=sheet_name, index=False)
                    worksheet = writer.sheets[sheet_name]
                    self._format_data_sheet(worksheet, safe_dataframe, report_key)
                    if report_key == "settings":
                        worksheet.sheet_state = "hidden"

                self._link_summary_to_detail(workbook)
                self._link_resistor_summary_to_detail(workbook)
                self._sync_summary_from_workspace(workbook)
                self._add_rd_decision_dropdowns(workbook)
                self._build_dashboard(overview, reports, report_views)

            os.replace(temporary_path, output_path)
        except Exception:
            temporary_path.unlink(missing_ok=True)
            raise

        return str(output_path)

    def _build_report_views(self, reports):
        return {
            "merge_candidates": self._merge_candidate_view(
                reports.get("merge_candidates", pd.DataFrame())
            ),
            "specification_summary": self._specification_summary_view(reports),
            "specification_detail": self._specification_detail_view(reports),
            "resistor_summary": self._resistor_summary_view(reports),
            "resistor_detail": self._resistor_detail_view(reports),
            "avl_candidate": self._avl_candidate_view(
                reports.get("avl_candidates", pd.DataFrame())
            ),
            "risk_review": self._risk_review_view(reports),
            "nearby_value": self._nearby_value_view(reports),
            "settings": self._settings_view(reports),
        }

    def _merge_candidate_view(self, merge_candidates):
        columns = [
            "Priority",
            "Merge Difficulty",
            "Difference",
            "Spec",
            "Merge PN",
            "Merge Qty",
            "Keep PN",
            "Keep Qty",
            "BOM Qty",
            "Reference Count",
            "Reason",
            "Action",
            "Score",
        ]
        if merge_candidates.empty:
            return pd.DataFrame(columns=columns)

        rows = []
        for _, candidate in merge_candidates.iterrows():
            difference = self._difference_label(candidate.get("Spec_Similarity", 0))
            rows.append(
                {
                    "Priority": candidate.get("Priority_Stars", "") or candidate.get("Priority", ""),
                    "Merge Difficulty": self._difficulty_label(candidate.get("Spec_Similarity", 0)),
                    "Difference": difference,
                    "Spec": self._compact_spec(candidate),
                    "Merge PN": candidate.get("Current_PN", ""),
                    "Merge Qty": candidate.get("Current_Qty", 0),
                    "Keep PN": candidate.get("Target_PN", ""),
                    "Keep Qty": candidate.get("Target_Qty", 0),
                    "BOM Qty": float(candidate.get("Current_Qty", 0) or 0) + float(candidate.get("Target_Qty", 0) or 0),
                    "Reference Count": self._reference_count(candidate.get("Item", "")),
                    "Reason": candidate.get("Reason", ""),
                    "Action": candidate.get("Recommendation", ""),
                    "Score": candidate.get("Merge_Score", 0),
                }
            )

        return pd.DataFrame(rows, columns=columns)

    def _specification_summary_view(self, reports):
        columns = ["Merge ID", "Review Item", "BOM Qty", "Keep Qty", "Merge Qty", "Priority", "Why Review", "RD Decision", "Detail"]
        rows = []
        merge_groups = self._merge_tree_groups(reports, component_type="C")
        assigned_parts = {}
        for group in merge_groups:
            self._assign_bucket_parts(assigned_parts, group["group_id"], [group["target"], *group["candidates"]])
            keep_quantity, merge_quantity = self._review_quantities(group["target"], group["candidates"])
            rows.append(
                {
                    "Merge ID": group["group_id"],
                    "Review Item": self._review_item(group["value"], group["spec_detail"]),
                    "BOM Qty": group["total_quantity"],
                    "Keep Qty": keep_quantity,
                    "Merge Qty": merge_quantity,
                    "Priority": group["potential"],
                    "Why Review": group["summary_reason"],
                    "RD Decision": "",
                    "Detail": "Open",
                }
            )

        for group in self._variant_summary_groups(reports, len(rows) + 1, component_type="C", assigned_parts=assigned_parts):
            rows.append(group)

        dataframe = pd.DataFrame(rows, columns=columns)
        if dataframe.empty:
            return dataframe
        dataframe["_Priority Rank"] = dataframe["Priority"].map(self._priority_rank)
        dataframe = dataframe.sort_values(
            ["_Priority Rank", "Merge Qty", "BOM Qty", "Review Item"],
            ascending=[False, False, False, True],
        ).drop(columns=["_Priority Rank"])
        return dataframe.reset_index(drop=True)

    def _specification_detail_view(self, reports):
        columns = [
            "Merge ID",
            "Keep PN",
            "Merge PN",
            "Keep Qty",
            "Merge Qty",
            "Difference",
            "Vendor",
            "Package",
            "Voltage",
            "Material",
            "RD Decision",
        ]
        rows = []
        assigned_parts = {}
        merge_groups = self._merge_tree_groups(reports, component_type="C")
        for group in merge_groups:
            self._assign_bucket_parts(assigned_parts, group["group_id"], [group["target"], *group["candidates"]])
            rows.extend(self._same_spec_review_card(group, columns))

        rows.extend(self._variant_detail_groups(reports, component_type="C", assigned_parts=assigned_parts, columns=columns, start_index=len(merge_groups) + 1))
        self._validate_unique_buckets(assigned_parts)

        rows = self._sort_detail_sections(rows, columns)
        return pd.DataFrame(rows, columns=columns)

    def _resistor_summary_view(self, reports):
        columns = ["Value", "PN Count", "Total Qty", "Action / Target PN", "Priority", "Reason (相似度分类)", "Detail", "Group", "RD Decision"]
        rows = []
        for group in self._resistor_value_groups(reports):
            rows.append(
                {
                    "Value": group["value"],
                    "PN Count": group["pn_count"],
                    "Total Qty": group["total_quantity"],
                    "Action / Target PN": group["target_action"],
                    "Priority": group["priority"],
                    "Reason (相似度分类)": group["reason"],
                    "Detail": "▶ Detail",
                    "Group": group["group_id"],
                    "RD Decision": "⏳ 待处理 (Pending)",
                }
            )
        return pd.DataFrame(rows, columns=columns)

    def _resistor_detail_view(self, reports):
        columns = ["Group", "Row Type", "Value", "Spec", "PN", "Qty", "Status", "Difference", "Why Listed"]
        rows = []
        for group in self._resistor_value_groups(reports):
            rows.append(
                {
                    "Group": group["group_id"],
                    "Row Type": "Value Header",
                    "Value": group["value"],
                    "Spec": f"{group['pn_count']} PN  |  Total {group['total_quantity']:g} pcs",
                    "PN": "",
                    "Qty": "",
                    "Status": "",
                    "Difference": group["nearby_values"],
                    "Why Listed": group["reason"],
                }
            )
            for spec in group["spec_groups"]:
                rows.append(
                    {
                        "Group": group["group_id"],
                        "Row Type": "Spec Header",
                        "Value": group["value"],
                        "Spec": spec["spec"],
                        "PN": "",
                        "Qty": spec["total_quantity"],
                        "Status": "Same Spec Group",
                        "Difference": "",
                        "Why Listed": spec["why_listed"],
                    }
                )
                for part in spec["parts"]:
                    rows.append(
                        {
                            "Group": group["group_id"],
                            "Row Type": "PN",
                            "Value": group["value"],
                            "Spec": spec["spec"],
                            "PN": part["part_number"],
                            "Qty": part["quantity"],
                            "Status": part["status"],
                            "Difference": part["difference"],
                            "Why Listed": spec["why_listed"],
                        }
                    )
            rows.append({column: "" for column in columns})
        return pd.DataFrame(rows, columns=columns)

    def _nearby_value_view(self, reports):
        columns = ["Current Value", "Current BOM Qty", "Nearby Value", "Candidate Qty", "Difference", "Tolerance Band", "Family", "Candidate PNs"]
        near_resistance = reports.get("near_resistance", pd.DataFrame())
        if near_resistance.empty:
            return pd.DataFrame(columns=columns)

        rows = []
        for _, pair in near_resistance.iterrows():
            value_a = pair.get("Value_A", "")
            value_b = pair.get("Value_B", "")
            difference = float(pair.get("Difference_Percent", 0) or 0)
            family = self._nearby_family(pair)
            rows.extend(
                self._nearby_reference_rows(
                    value_a,
                    pair.get("Quantity_A", 0),
                    value_b,
                    pair.get("Quantity_B", 0),
                    difference,
                    family,
                    pair.get("Part_Numbers_B", ""),
                    columns,
                )
            )
            reverse_difference = -difference / (1 + difference / 100) if difference != -100 else 0
            rows.extend(
                self._nearby_reference_rows(
                    value_b,
                    pair.get("Quantity_B", 0),
                    value_a,
                    pair.get("Quantity_A", 0),
                    reverse_difference,
                    family,
                    pair.get("Part_Numbers_A", ""),
                    columns,
                )
            )

        dataframe = pd.DataFrame(rows, columns=columns)
        if dataframe.empty:
            return dataframe
        dataframe["_Abs Difference"] = dataframe["Difference"].map(self._percent_abs_value)
        dataframe = dataframe.sort_values(
            ["Current Value", "_Abs Difference", "Candidate Qty", "Nearby Value"],
            ascending=[True, True, False, True],
        ).drop(columns=["_Abs Difference"])
        return dataframe.reset_index(drop=True)

    @staticmethod
    def _avl_candidate_view(avl_candidates):
        columns = ["Spec", "Preferred PN", "PN Count", "Vendor", "Qty", "Status", "Suggestion"]
        if avl_candidates.empty:
            return pd.DataFrame(columns=columns)
        dataframe = avl_candidates.rename(
            columns={
                "Normalize_Key": "Spec",
                "Preferred_PN": "Preferred PN",
                "Part_Number_Count": "PN Count",
                "Vendors": "Vendor",
                "Total_Quantity": "Qty",
                "AVL_Readiness": "Status",
                "Recommendation": "Suggestion",
            }
        )
        return dataframe.reindex(columns=columns)

    def _risk_review_view(self, reports):
        columns = ["Severity", "Risk", "PN", "Reference Count", "Finding", "Suggestion"]
        risk_components = reports.get("risk_components", pd.DataFrame())
        rows = []
        for _, risk in risk_components.iterrows():
            rows.append(
                {
                    "Severity": risk.get("Severity", ""),
                    "Risk": risk.get("Risk_Category", ""),
                    "PN": risk.get("Part_Number", ""),
                    "Reference Count": self._reference_count(risk.get("Reference", "")),
                    "Finding": risk.get("Finding", ""),
                    "Suggestion": risk.get("Recommendation", ""),
                }
            )
        return pd.DataFrame(rows, columns=columns)

    def _merge_tree_groups(self, reports, component_type=None):
        normalized = reports.get("normalized_bom", pd.DataFrame())
        required_columns = {"Component_Type", "Normalize_Key", "Part_Number", "Quantity_Normalized"}
        if normalized.empty or not required_columns.issubset(normalized.columns):
            return []

        candidates = normalized[
            normalized["Component_Type"].isin(["C", "R"])
            & normalized["Normalize_Key"].fillna("").ne("")
            & normalized["Part_Number"].fillna("").astype(str).str.strip().ne("")
            & ~normalized.get("Is_Second_Source", pd.Series(False, index=normalized.index))
        ]
        if component_type:
            candidates = candidates[candidates["Component_Type"].eq(component_type)]
        groups = []
        for _, group in candidates.groupby(["Component_Type", "Normalize_Key"], sort=True):
            part_groups = self._part_groups(group)
            if len(part_groups) < 2:
                continue

            target = max(
                part_groups,
                key=lambda item: (item["quantity"], item["line_count"], item["part_number"]),
            )
            candidate_parts = [part for part in part_groups if part["part_number"] != target["part_number"]]
            candidate_parts.sort(key=lambda item: (item["quantity"], item["part_number"]))
            potential = self._group_potential(candidate_parts)
            value, spec_detail = self._split_spec(group.iloc[0])
            groups.append(
                {
                    "value": value,
                    "spec_detail": spec_detail,
                    "spec": self._compact_spec(group.iloc[0]),
                    "pn_count": len(part_groups),
                    "total_quantity": float(group["Quantity_Normalized"].sum()),
                    "target": target,
                    "candidates": candidate_parts,
                    "potential": potential,
                    "summary_reason": self._group_summary_reason(candidate_parts, target),
                }
            )

        groups.sort(key=lambda item: (item["potential"], item["total_quantity"]), reverse=True)
        for index, group in enumerate(groups, start=1):
            group["group_id"] = f"M-{index:03d}"
        return groups

    def _resistor_value_groups(self, reports):
        normalized = reports.get("normalized_bom", pd.DataFrame())
        required_columns = {"Component_Type", "Normalized_Value", "Normalize_Key", "Part_Number", "Quantity_Normalized"}
        if normalized.empty or not required_columns.issubset(normalized.columns):
            return []

        candidates = normalized[
            normalized["Component_Type"].eq("R")
            & normalized["Normalized_Value"].fillna("").ne("")
            & normalized["Normalize_Key"].fillna("").ne("")
            & normalized["Part_Number"].fillna("").astype(str).str.strip().ne("")
            & ~normalized.get("Is_Second_Source", pd.Series(False, index=normalized.index))
        ].copy()
        if candidates.empty:
            return []

        nearby_lookup = self._nearby_values_by_resistance(reports.get("near_resistance", pd.DataFrame()))
        groups = []
        for value, value_group in candidates.groupby("Normalized_Value", sort=True):
            spec_groups = self._resistor_spec_groups(value_group)
            part_numbers = self._unique_text(value_group["Part_Number"])
            total_quantity = float(value_group["Quantity_Normalized"].sum())
            nearby_candidates = nearby_lookup.get(str(value), [])
            has_high_frequency_nearby = any(
                abs(candidate["difference_percent"]) <= 3 and candidate["quantity"] > total_quantity
                for candidate in nearby_candidates
            )
            if len(part_numbers) < 2 and not has_high_frequency_nearby:
                continue
            classification = self._resistor_classification(value_group, spec_groups, nearby_candidates)
            groups.append(
                {
                    "value": value,
                    "pn_count": len(part_numbers),
                    "total_quantity": total_quantity,
                    "nearby_values": ", ".join(candidate["label"] for candidate in nearby_candidates[:3]),
                    "target_action": classification["target_action"],
                    "priority": classification["priority"],
                    "reason": classification["reason"],
                    "spec_groups": spec_groups,
                }
            )

        groups.sort(key=lambda item: (item["total_quantity"], item["value"]))
        for index, group in enumerate(groups, start=1):
            group["group_id"] = f"R{index:03d}"
        return groups

    def _resistor_spec_groups(self, value_group):
        spec_groups = []
        for _, spec_group in value_group.groupby("Normalize_Key", sort=True):
            part_groups = self._part_groups(spec_group)
            target_part_number = max(
                part_groups,
                key=lambda item: (item["quantity"], item["line_count"], item["part_number"]),
            )["part_number"]
            part_groups.sort(key=lambda item: (item["quantity"], item["part_number"]))
            representative = spec_group.iloc[0]
            parts = []
            for part in part_groups:
                difference, _ = self._part_difference(part["representative"], representative)
                parts.append(
                    {
                        "part_number": part["part_number"],
                        "quantity": part["quantity"],
                        "status": "Target" if part["part_number"] == target_part_number else "Review",
                        "difference": "" if difference == "🟢 Same spec" else difference,
                    }
                )
            spec_groups.append(
                {
                    "spec": self._compact_spec(representative),
                    "total_quantity": float(spec_group["Quantity_Normalized"].sum()),
                    "why_listed": self._spec_group_why_listed(spec_group, part_groups),
                    "parts": parts,
                }
            )
        spec_groups.sort(key=lambda item: (item["total_quantity"], item["spec"]))
        return spec_groups

    def _resistor_classification(self, value_group, spec_groups, nearby_candidates):
        package_values = self._unique_text(value_group["Package_Identity"])
        tolerance_values = self._unique_text(value_group["Tolerance"])
        if len(package_values) > 1:
            return {
                "target_action": "Review Required",
                "priority": "★★★",
                "reason": f"🟡 Package {self._compact_value_pair(package_values)}",
            }
        if len(tolerance_values) > 1:
            return {
                "target_action": "Review Required",
                "priority": "★★★",
                "reason": f"🟡 Tolerance {self._compact_value_pair(tolerance_values)}",
            }

        near_candidates = [candidate for candidate in nearby_candidates if abs(candidate["difference_percent"]) <= 3]
        if near_candidates:
            candidate = near_candidates[0]
            return {
                "target_action": "Review Required",
                "priority": "★★",
                "reason": (
                    f"🟠 Near Value {self._display_resistance_value(value_group.iloc[0]['Normalized_Value'])} "
                    f"➔ {self._display_resistance_value(candidate['value'])} "
                    f"({self._signed_percent(candidate['difference_percent'])})"
                ),
            }

        part_groups = [part for spec in spec_groups for part in spec["parts"]]
        if len(part_groups) > 1:
            target = max(part_groups, key=lambda item: (item["quantity"], item["part_number"]))
            return {
                "target_action": target["part_number"],
                "priority": "★★★★★",
                "reason": "🟢 Same Spec",
            }

        return {
            "target_action": "Review Required",
            "priority": "★",
            "reason": "Review",
        }

    def _spec_group_why_listed(self, spec_group, part_groups):
        if len(part_groups) > 1:
            return f"{len(part_groups)} PN"
        total_quantity = float(spec_group["Quantity_Normalized"].sum())
        if total_quantity <= 2:
            return f"Qty={total_quantity:g}"
        vendor_count = len(self._unique_text(spec_group.get("Vendor", pd.Series(dtype=object)), case_insensitive=True))
        if vendor_count > 1:
            return "Vendor Mixed"
        return "Same Spec"

    def _nearby_values_by_resistance(self, near_resistance):
        nearby = {}
        if near_resistance.empty:
            return nearby
        for _, pair in near_resistance.iterrows():
            value_a = str(pair.get("Value_A", "") or "")
            value_b = str(pair.get("Value_B", "") or "")
            difference = float(pair.get("Difference_Percent", 0) or 0)
            quantity_a = float(pair.get("Quantity_A", 0) or 0)
            quantity_b = float(pair.get("Quantity_B", 0) or 0)
            if value_a and value_b:
                value_a_to_b = -difference / (1 + difference / 100) if difference != -100 else 0
                nearby.setdefault(value_a, []).append(
                    {
                        "value": value_b,
                        "difference_percent": value_a_to_b,
                        "quantity": quantity_b,
                        "label": f"{self._display_resistance_value(value_b)} ({self._signed_percent(value_a_to_b)})",
                    }
                )
                nearby.setdefault(value_b, []).append(
                    {
                        "value": value_a,
                        "difference_percent": difference,
                        "quantity": quantity_a,
                        "label": f"{self._display_resistance_value(value_a)} ({self._signed_percent(difference)})",
                    }
                )
        return {
            value: sorted(
                [candidate for candidate in values if abs(candidate["difference_percent"]) <= 5],
                key=lambda candidate: (-candidate["quantity"], abs(candidate["difference_percent"]), candidate["value"]),
            )[:3]
            for value, values in nearby.items()
        }

    def _part_groups(self, group):
        part_groups = []
        for part_number, part_group in group.groupby("Part_Number", sort=False):
            representative = part_group.iloc[0]
            part_groups.append(
                {
                    "part_number": str(part_number),
                    "quantity": float(part_group["Quantity_Normalized"].sum()),
                    "line_count": len(part_group),
                    "vendor": ", ".join(self._unique_text(part_group.get("Vendor", pd.Series(dtype=object)), case_insensitive=True)),
                    "representative": representative,
                }
            )
        return part_groups

    @staticmethod
    def _group_potential(candidate_parts):
        if not candidate_parts:
            return ""
        minimum_quantity = min(part["quantity"] for part in candidate_parts)
        if minimum_quantity <= 2:
            return "★★★★★"
        if minimum_quantity <= 5:
            return "★★★★"
        if minimum_quantity <= 10:
            return "★★★"
        return "★★"

    def _group_summary_reason(self, candidate_parts, target):
        if not candidate_parts:
            return "No merge candidate"
        differences = []
        for candidate in candidate_parts:
            difference, _ = self._part_difference(candidate["representative"], target["representative"])
            if difference and difference not in differences:
                differences.append(difference)
        if differences == ["🟢 Same spec"]:
            return "Same Spec"
        if differences == ["🟢 Vendor"]:
            return "Vendor Only"
        return " / ".join(differences) or "Same Spec"

    def _variant_summary_groups(self, reports, start_index, component_type=None, assigned_parts=None):
        rows = []
        assigned_parts = assigned_parts if assigned_parts is not None else {}
        normalized = reports.get("normalized_bom", pd.DataFrame())
        group_index = start_index
        for report_key, status, label, attribute in (
            ("different_package", "★★★", "🟡 Package", "Package_Identity"),
            ("different_voltage", "★★", "🟠 Voltage", "Voltage"),
            ("different_material", "★", "🔴 Material", "Dielectric"),
        ):
            variants = reports.get(report_key, pd.DataFrame())
            if component_type and "Component_Type" in variants:
                variants = variants[variants["Component_Type"].eq(component_type)]
            for _, variant in variants.iterrows():
                values = str(variant.get("Attribute_Values", ""))
                members = self._variant_member_parts(normalized, variant, attribute, assigned_parts)
                if len(members) < 2:
                    continue
                target = self._variant_target_member(members, attribute) if members else None
                review_members = [
                    member
                    for member in members
                    if target and member["part_number"] != target["part_number"]
                    and self._engineering_difference(member["representative"], target["representative"])[0]
                ]
                if not target or not review_members:
                    continue
                reason = self._variant_reason(members, target, attribute, label) if target else f"{label} {self._compact_attribute_values(values)}"
                keep_quantity = target["quantity"] if target else 0
                merge_quantity = sum(member["quantity"] for member in review_members)
                total_quantity = keep_quantity + merge_quantity
                group_id = f"M-{group_index:03d}"
                self._assign_bucket_parts(assigned_parts, group_id, [target, *review_members])
                rows.append(
                    {
                        "Merge ID": group_id,
                        "Review Item": self._review_item(variant.get("Normalized_Value", ""), self._compact_attribute_values(values)),
                        "BOM Qty": total_quantity or variant.get("Total_Quantity", 0),
                        "Keep Qty": keep_quantity,
                        "Merge Qty": merge_quantity,
                        "Priority": status,
                        "Why Review": reason,
                        "RD Decision": "",
                        "Detail": "Open",
                    }
                )
                group_index += 1
        return rows

    def _variant_detail_groups(self, reports, component_type=None, assigned_parts=None, columns=None, start_index=1):
        rows = []
        assigned_parts = assigned_parts if assigned_parts is not None else {}
        columns = columns or [
            "Merge ID",
            "Keep PN",
            "Merge PN",
            "Keep Qty",
            "Merge Qty",
            "Difference",
            "Vendor",
            "Package",
            "Voltage",
            "Material",
            "RD Decision",
        ]
        group_index = start_index
        normalized = reports.get("normalized_bom", pd.DataFrame())
        for report_key, label, attribute in (
            ("different_package", "🟡 Package", "Package_Identity"),
            ("different_voltage", "🟠 Voltage", "Voltage"),
            ("different_material", "🔴 Material", "Dielectric"),
        ):
            variants = reports.get(report_key, pd.DataFrame())
            if component_type and "Component_Type" in variants:
                variants = variants[variants["Component_Type"].eq(component_type)]
            for _, variant in variants.iterrows():
                group_id = f"M-{group_index:03d}"
                members = self._variant_member_parts(normalized, variant, attribute, assigned_parts)
                if len(members) < 2:
                    continue
                total_quantity = sum(member["quantity"] for member in members)
                feature_distribution = self._feature_distribution(members, attribute, total_quantity)
                target = self._variant_target_member(members, attribute)
                review_members = [
                    member
                    for member in members
                    if member["part_number"] != target["part_number"]
                    and self._engineering_difference(member["representative"], target["representative"])[0]
                ]
                if not review_members:
                    continue
                self._assign_bucket_parts(assigned_parts, group_id, [target, *review_members])
                primary_difference = self._primary_difference_type(review_members, target)
                rows.extend(
                    self._review_required_card(
                        group_id,
                        variant.get("Normalized_Value", ""),
                        primary_difference,
                        target,
                        review_members,
                        feature_distribution,
                        total_quantity,
                        columns,
                    )
                )
                rows.append({column: "" for column in columns})
                group_index += 1
        return rows

    def _same_spec_review_card(self, group, columns):
        """Build a KEEP/MERGE worksheet block for exact-spec PN consolidation."""
        target = group["target"]
        candidates = group["candidates"]
        rows = [self._workspace_row(group["group_id"], target, None, difference="KEEP")]
        for candidate in sorted(candidates, key=lambda item: (-item["quantity"], item["part_number"])):
            _, difference = self._engineering_difference(candidate["representative"], target["representative"])
            rows.append(self._workspace_row(group["group_id"], target, candidate, difference or "Same Spec"))
        rows.append({column: "" for column in columns})
        return rows

    def _review_required_card(self, group_id, value, difference_type, target, review_members, feature_distribution, total_quantity, columns):
        """Build a KEEP/MERGE worksheet block for one review-required group."""
        rows = [self._workspace_row(group_id, target, None, difference="KEEP")]
        for member in sorted(review_members, key=lambda item: (-item["quantity"], item["part_number"])):
            _, difference = self._engineering_difference(member["representative"], target["representative"])
            rows.append(self._workspace_row(group_id, target, member, difference))
        return rows

    def _workspace_row(self, group_id, keep_part, merge_part=None, difference=""):
        """Return one normalized Merge Workspace row using KEEP/MERGE language."""
        keep_row = keep_part["representative"]
        merge_row = merge_part["representative"] if merge_part else keep_row
        return {
            "Merge ID": group_id,
            "Keep PN": keep_part["part_number"],
            "Merge PN": merge_part["part_number"] if merge_part else "",
            "Keep Qty": keep_part["quantity"],
            "Merge Qty": merge_part["quantity"] if merge_part else "",
            "Difference": difference,
            "Vendor": self._direction_text(merge_row.get("Vendor", ""), keep_row.get("Vendor", "")) if merge_part else keep_row.get("Vendor", ""),
            "Package": self._direction_text(merge_row.get("Package_Identity", ""), keep_row.get("Package_Identity", "")) if merge_part else keep_row.get("Package_Identity", ""),
            "Voltage": self._direction_text(merge_row.get("Voltage", ""), keep_row.get("Voltage", "")) if merge_part else keep_row.get("Voltage", ""),
            "Material": self._direction_text(merge_row.get("Dielectric", ""), keep_row.get("Dielectric", "")) if merge_part else keep_row.get("Dielectric", ""),
            "RD Decision": "" if merge_part else "",
        }

    @staticmethod
    def _review_quantities(keep_part, merge_parts):
        """Return the three report quantities as Keep Qty and Merge Qty."""
        keep_quantity = keep_part["quantity"] if keep_part else 0
        merge_quantity = sum(part["quantity"] for part in merge_parts)
        return keep_quantity, merge_quantity

    def _nearby_reference_rows(self, current_value, current_quantity, nearby_value, candidate_quantity, difference, family, candidate_pns, columns):
        """Return a nearby resistor row only when the candidate is within an RD-useful band."""
        absolute_difference = abs(float(difference or 0))
        if absolute_difference > 10:
            return []
        return [
            {
                "Current Value": current_value,
                "Current BOM Qty": float(current_quantity or 0),
                "Nearby Value": nearby_value,
                "Candidate Qty": float(candidate_quantity or 0),
                "Difference": self._signed_percent(difference),
                "Tolerance Band": self._tolerance_band(absolute_difference),
                "Family": family,
                "Candidate PNs": candidate_pns,
            }
        ]

    @staticmethod
    def _tolerance_band(absolute_difference):
        """Classify nearby resistor values into the review bands RD asked for."""
        if absolute_difference <= 2:
            return "±2%"
        if absolute_difference <= 5:
            return "±5%"
        return "±10%"

    @staticmethod
    def _percent_abs_value(percent_text):
        """Convert a signed percent string into an absolute numeric sort value."""
        try:
            return abs(float(str(percent_text).replace("%", "")))
        except (TypeError, ValueError):
            return 999

    @staticmethod
    def _review_item(value, detail):
        """Combine value and spec detail into the Summary review item label."""
        return f"{value} | {detail}" if detail else str(value or "")

    @staticmethod
    def _direction_text(current, recommended):
        """Show Current -> Recommended when values differ, otherwise show the stable value."""
        current_text = str(current or "").strip()
        recommended_text = str(recommended or "").strip()
        if current_text and recommended_text and current_text != recommended_text:
            return f"{current_text} → {recommended_text}"
        return recommended_text or current_text

    @staticmethod
    def _priority_rank(priority):
        """Convert star priority text into a sortable numeric rank."""
        return str(priority or "").count("★")

    def _feature_from_distribution(self, distribution, target):
        """Pick the highest-quantity feature already selected as the keep side."""
        target_features = {self._row_feature(target["representative"], field) for field in ("Package_Identity", "Voltage", "Dielectric", "Tolerance")}
        for item in distribution:
            if item["feature"] in target_features:
                return item["feature"]
        return distribution[0]["feature"] if distribution else "Keep"

    @staticmethod
    def _card_current_value(difference, difference_type):
        """Extract a short Current value for the card row from a multi-line difference."""
        for block in str(difference or "").split("\n\n"):
            lines = block.splitlines()
            if lines and lines[0] == difference_type and "→" in lines[-1]:
                return lines[-1].split("→", 1)[0].strip()
        return str(difference or "")

    @staticmethod
    def _card_recommendation(difference):
        """Extract the recommended value from the first visible difference."""
        for block in str(difference or "").split("\n\n"):
            lines = block.splitlines()
            if lines and "→" in lines[-1]:
                return lines[-1].split("→", 1)[1].strip()
        return "Review"

    def _variant_member_parts(self, normalized, variant, attribute, assigned_parts):
        required_columns = {"Component_Type", "Part_Number", "Quantity_Normalized", attribute}
        if normalized.empty or not required_columns.issubset(normalized.columns):
            return []

        candidates = normalized[
            normalized["Component_Type"].eq(variant.get("Component_Type", ""))
            & normalized["Normalized_Value"].eq(variant.get("Normalized_Value", ""))
            & normalized[attribute].fillna("").astype(str).str.strip().ne("")
            & normalized["Part_Number"].fillna("").astype(str).str.strip().ne("")
            & normalized.get("Comparison_Eligible", pd.Series(True, index=normalized.index))
            & ~normalized.get("Is_Second_Source", pd.Series(False, index=normalized.index))
        ].copy()

        for field, value in self._parse_variant_context(variant.get("Specification_Context", "")).items():
            if field in candidates.columns:
                candidates = candidates[candidates[field].fillna("").astype(str).eq(value)]

        attribute_values = {value.strip() for value in str(variant.get("Attribute_Values", "")).split(",") if value.strip()}
        if attribute_values:
            candidates = candidates[candidates[attribute].fillna("").astype(str).str.strip().isin(attribute_values)]

        parts = self._part_groups(candidates)
        return [part for part in parts if part["part_number"] not in assigned_parts]

    def _variant_target_member(self, members, attribute):
        if not members:
            return None
        totals = {}
        for member in members:
            feature = self._row_feature(member["representative"], attribute)
            totals[feature] = totals.get(feature, 0) + member["quantity"]
        target_feature = max(totals.items(), key=lambda item: (item[1], item[0]))[0]
        target_members = [member for member in members if self._row_feature(member["representative"], attribute) == target_feature]
        return max(target_members, key=lambda item: (item["quantity"], item["part_number"]))

    @staticmethod
    def _parse_variant_context(context):
        values = {}
        for segment in str(context or "").split(";"):
            if "=" not in segment:
                continue
            field, value = segment.split("=", 1)
            values[field.strip()] = "" if value.strip() == "(blank)" else value.strip()
        return values

    def _feature_distribution(self, members, attribute, total_quantity):
        quantities = {}
        for member in members:
            feature = self._row_feature(member["representative"], attribute)
            quantities[feature] = quantities.get(feature, 0) + member["quantity"]
        distribution = [
            {"feature": feature, "quantity": quantity, "share": self._quantity_share(quantity, total_quantity)}
            for feature, quantity in quantities.items()
        ]
        return sorted(distribution, key=lambda item: (-item["quantity"], item["feature"]))

    @staticmethod
    def _format_feature_distribution(distribution):
        return " | ".join(f"{item['feature']}: {item['quantity']:g} pcs ({str(item['share']).split('(', 1)[-1]}" for item in distribution)

    def _engineering_difference(self, current, target):
        """Return only engineering-relevant fields that differ from the target PN."""
        checks = (
            ("Vendor", "Vendor", "Vendor"),
            ("Package_Identity", "Package", "Package"),
            ("Voltage", "Voltage", "Voltage"),
            ("Dielectric", "Material", "Material"),
            ("Tolerance", "Tolerance", "Tolerance"),
        )
        difference_types = []
        difference_lines = []
        for field, difference_type, label in checks:
            current_value = str(current.get(field, "") or "")
            target_value = str(target.get(field, "") or "")
            if current_value != target_value:
                difference_types.append(difference_type)
                difference_lines.append(f"{label}\n{current_value or '(blank)'} → {target_value or '(blank)'}")

        return " / ".join(difference_types), "\n\n".join(difference_lines)

    def _group_primary_difference(self, candidates, target):
        """Find the first visible difference type for sorting and scanning a merge group."""
        return self._primary_difference_type(candidates, target) or "Same Spec"

    def _primary_difference_type(self, members, target):
        """Sort review sections by RD scan order: Vendor, Package, Voltage, Material."""
        order = {"Vendor": 0, "Package": 1, "Voltage": 2, "Material": 3, "Tolerance": 4}
        found = []
        for member in members:
            difference_type, _ = self._engineering_difference(member["representative"], target["representative"])
            found.extend(part.strip() for part in difference_type.split("/") if part.strip())
        if not found:
            return ""
        return sorted(set(found), key=lambda item: order.get(item, 99))[0]

    def _sort_detail_sections(self, rows, columns):
        """Keep each Merge Workspace group intact while preserving Merge ID order."""
        sections = []
        section = []
        for row in rows:
            if not any(row.values()):
                if section:
                    sections.append(section)
                    section = []
                sections.append([row])
            elif row.get("Merge ID") and (not section or row.get("Merge ID") != section[0].get("Merge ID")):
                if section:
                    sections.append(section)
                section = [row]
            elif section:
                section.append(row)
            else:
                sections.append([row])
        if section:
            sections.append(section)

        def section_key(section_rows):
            merge_id = str(section_rows[0].get("Merge ID", ""))
            return merge_id or "ZZZ"

        sorted_rows = []
        for section_rows in sorted(sections, key=section_key):
            sorted_rows.extend(section_rows)
        return [row if row else {column: "" for column in columns} for row in sorted_rows]

    @staticmethod
    def _quantity_percent(quantity, total_quantity):
        """Format a compact integer percentage for group summary rows."""
        if not total_quantity:
            return "0%"
        return f"{round(float(quantity) / float(total_quantity) * 100):g}%"

    @staticmethod
    def _row_feature(row, attribute):
        return str(row.get(attribute, "") or "(blank)").strip() or "(blank)"

    def _variant_reason(self, members, target, attribute, label):
        target_feature = self._row_feature(target["representative"], attribute)
        source_features = []
        for member in members:
            feature = self._row_feature(member["representative"], attribute)
            if feature != target_feature and feature not in source_features:
                source_features.append(feature)
        feature_text = ", ".join(source_features) if source_features else target_feature
        reason = f"{label} {feature_text} ➔ {target_feature}"

        extra_differences = []
        for member in members:
            difference, _ = self._part_differences(member["representative"], target["representative"], excluded_fields={attribute})
            if difference not in {"🟢 Same spec", "🟢 Vendor"} and difference not in extra_differences:
                extra_differences.append(difference)
        if extra_differences:
            reason = f"{reason} / {' / '.join(extra_differences)}"
        return reason

    @staticmethod
    def _assign_bucket_parts(assigned_parts, group_id, parts):
        for part in parts:
            part_number = part.get("part_number", "")
            if not part_number:
                continue
            existing_group = assigned_parts.get(part_number)
            if existing_group and existing_group != group_id:
                raise ValueError(f"Part number {part_number} appears in multiple review buckets: {existing_group}, {group_id}")
            assigned_parts[part_number] = group_id

    @staticmethod
    def _validate_unique_buckets(assigned_parts):
        seen = {}
        for part_number, group_id in assigned_parts.items():
            previous = seen.get(part_number)
            if previous and previous != group_id:
                raise ValueError(f"Part number {part_number} appears in multiple review buckets: {previous}, {group_id}")
            seen[part_number] = group_id

    def _part_difference(self, current, target):
        checks = (
            ("Normalized_Value", "🔴 Value", "Value"),
            ("Dielectric", "🔴 Material", "Material"),
            ("Package_Identity", "🟡 Package", "Package"),
            ("Voltage", "🟠 Voltage", "Voltage"),
            ("Tolerance", "🟡 Tolerance", "Tolerance"),
        )
        for field, label, name in checks:
            current_value = str(current.get(field, "") or "")
            target_value = str(target.get(field, "") or "")
            if current_value != target_value:
                return label, f"{name} {current_value or '(blank)'} -> {target_value or '(blank)'}"

        current_vendor = str(current.get("Vendor", "") or "").casefold()
        target_vendor = str(target.get("Vendor", "") or "").casefold()
        if current_vendor != target_vendor:
            return "🟢 Vendor", "Vendor different"
        return "🟢 Same spec", "Same specification"

    def _part_differences(self, current, target, excluded_fields=None):
        excluded_fields = excluded_fields or set()
        checks = (
            ("Normalized_Value", "🔴 Value", "Value"),
            ("Dielectric", "🔴 Material", "Material"),
            ("Package_Identity", "🟡 Package", "Package"),
            ("Voltage", "🟠 Voltage", "Voltage"),
            ("Tolerance", "🟡 Tolerance", "Tolerance"),
        )
        differences = []
        reasons = []
        for field, label, name in checks:
            if field in excluded_fields:
                continue
            current_value = str(current.get(field, "") or "")
            target_value = str(target.get(field, "") or "")
            if current_value != target_value:
                differences.append(f"{label} {current_value or '(blank)'} ➔ {target_value or '(blank)'}")
                reasons.append(f"{name} {current_value or '(blank)'} -> {target_value or '(blank)'}")

        if differences:
            return " / ".join(differences), "; ".join(reasons)

        current_vendor = str(current.get("Vendor", "") or "").casefold()
        target_vendor = str(target.get("Vendor", "") or "").casefold()
        if current_vendor != target_vendor:
            return "🟢 Vendor", "Vendor different"
        return "🟢 Same spec", "Same specification"

    def _settings_view(self, reports):
        rows = []
        for section, dataframe in (
            ("Summary", reports.get("summary", pd.DataFrame())),
            ("Metadata", reports.get("report_metadata", pd.DataFrame())),
            ("Rules", reports.get("rule_library", pd.DataFrame())),
        ):
            if dataframe.empty:
                continue
            for _, row in dataframe.iterrows():
                rows.append(
                    {
                        "Section": section,
                        "Name": row.get("Metric", row.get("Property", row.get("Rule_Path", ""))),
                        "Value": row.get("Value", ""),
                        "Source": row.get("Source", ""),
                    }
                )
        return pd.DataFrame(rows, columns=["Section", "Name", "Value", "Source"])

    def _build_dashboard(self, worksheet, reports, report_views):
        worksheet.sheet_view.showGridLines = False
        worksheet.freeze_panes = "A17"
        worksheet.sheet_view.zoomScale = 90
        worksheet.merge_cells("A1:H2")
        title = worksheet["A1"]
        title.value = "BOM Intelligence Overview"
        title.fill = PatternFill("solid", fgColor=self.COLORS["navy"])
        title.font = Font(name="Aptos Display", size=24, bold=True, color=self.COLORS["white"])
        title.alignment = Alignment(vertical="center")

        for row in worksheet[1:2]:
            for cell in row:
                cell.fill = PatternFill("solid", fgColor=self.COLORS["navy"])

        worksheet.merge_cells("A3:H3")
        worksheet["A3"] = "Global health check: Dashboard shows scope; Summary queues review; RD records decisions in Merge Workspace."
        worksheet["A3"].font = Font(name="Aptos", size=11, color=self.COLORS["muted"])
        worksheet["A3"].alignment = Alignment(vertical="center")

        summary = self._metric_lookup(reports.get("summary", pd.DataFrame()))
        normalized = reports.get("normalized_bom", pd.DataFrame())
        merge_view = report_views.get("merge_candidates", pd.DataFrame())
        merge_raw = reports.get("merge_candidates", pd.DataFrame())
        capacitor_summary = report_views.get("specification_summary", pd.DataFrame())
        resistor_summary = report_views.get("resistor_summary", pd.DataFrame())
        project = self._project_name(reports)
        valid_part_numbers = normalized["Part_Number"].replace("", pd.NA) if "Part_Number" in normalized else pd.Series(dtype=object)
        total_pn = int(valid_part_numbers.nunique()) if not valid_part_numbers.empty else 0
        capacitor_pn = self._component_part_count(normalized, "C")
        resistor_pn = self._component_part_count(normalized, "R")
        merge_pn_reduction = int(merge_raw["Current_PN"].replace("", pd.NA).nunique()) if "Current_PN" in merge_raw else 0
        cards = (
            ("Project", project, "Source BOM"),
            ("Total PN", total_pn, f"{summary.get('BOM Lines', 0)} BOM lines"),
            ("Capacitor PN", capacitor_pn, "Unique capacitor PNs"),
            ("Resistor PN", resistor_pn, "Unique resistor PNs"),
            ("Capacitor Review", len(capacitor_summary), "similarity groups"),
            ("Resistor Review", len(resistor_summary), "similarity groups"),
        )
        card_ranges = (
            "A5:B8",
            "C5:D8",
            "E5:F8",
            "A9:B12",
            "C9:D12",
            "E9:F12",
        )
        for card, cell_range in zip(cards, card_ranges):
            self._draw_kpi(worksheet, cell_range, *card)

        worksheet.merge_cells("G5:H8")
        benefit_cell = worksheet["G5"]
        benefit_cell.value = f"Estimated Benefit\n\nIf all accepted, reduce up to {merge_pn_reduction} redundant PN(s)."
        benefit_cell.fill = PatternFill("solid", fgColor=self.COLORS["cyan"])
        benefit_cell.font = Font(name="Aptos", size=12, bold=True, color=self.COLORS["ink"])
        benefit_cell.alignment = Alignment(vertical="center", wrap_text=True)
        benefit_cell.border = self._thin_border()

        worksheet.merge_cells("G9:H12")
        conclusion_cell = worksheet["G9"]
        conclusion_cell.value = (
            f"AI Review Summary\n\n"
            f"Found {len(capacitor_summary)} capacitor group(s) for RD review.\n"
            f"Found {len(resistor_summary)} resistor group(s) for RD review."
        )
        conclusion_cell.fill = PatternFill("solid", fgColor=self.COLORS["white"])
        conclusion_cell.font = Font(name="Aptos", size=11, bold=True, color=self.COLORS["ink"])
        conclusion_cell.alignment = Alignment(vertical="center", wrap_text=True)
        conclusion_cell.border = self._thin_border()

        self._write_top_merge_table(worksheet, merge_view)

        metadata = self._metric_lookup(
            reports.get("report_metadata", pd.DataFrame()),
            key_column="Property",
        )
        worksheet.merge_cells("A44:H44")
        worksheet["A44"] = (
            f"Generated: {metadata.get('Generated UTC', '')}  |  "
            f"Rules: {metadata.get('Rule Source', '')}  |  "
            f"Platform: {metadata.get('Platform Version', '')}"
        )
        worksheet["A44"].font = Font(name="Aptos", size=9, color=self.COLORS["muted"])
        worksheet["A44"].alignment = Alignment(vertical="center", wrap_text=True)

        for column in "ABCDEFGH":
            worksheet.column_dimensions[column].width = 18
        worksheet.column_dimensions["D"].width = 24
        worksheet.column_dimensions["E"].width = 24
        worksheet.column_dimensions["F"].width = 12

    def _write_top_merge_table(self, worksheet, merge_view):
        worksheet.merge_cells("A15:H15")
        worksheet["A15"] = "★★★★★ Priority 1"
        worksheet["A15"].font = Font(name="Aptos Display", size=16, bold=True, color=self.COLORS["teal"])
        worksheet["A15"].alignment = Alignment(vertical="center")

        headers = ["Spec", "Merge PN", "Merge Qty", "Keep PN", "Keep Qty", "Difference", "Difficulty", "Action"]
        for column_index, header in enumerate(headers, start=1):
            cell = worksheet.cell(16, column_index, header)
            cell.fill = PatternFill("solid", fgColor=self.COLORS["navy"])
            cell.font = Font(name="Aptos", size=10, bold=True, color=self.COLORS["white"])
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        if merge_view.empty:
            worksheet.cell(17, 1, "No Priority 1 merge candidate found.")
            return

        top_rows = merge_view[merge_view["Priority"].eq("★★★★★")].head(15)
        if top_rows.empty:
            top_rows = merge_view.head(15)

        for row_index, (_, row) in enumerate(top_rows.iterrows(), start=17):
            values = [
                row.get("Spec", ""),
                row.get("Merge PN", ""),
                row.get("Merge Qty", ""),
                row.get("Keep PN", ""),
                row.get("Keep Qty", ""),
                row.get("Difference", ""),
                row.get("Merge Difficulty", ""),
                row.get("Action", ""),
            ]
            for column_index, value in enumerate(values, start=1):
                cell = worksheet.cell(row_index, column_index, self._safe_excel_value(value))
                cell.font = Font(name="Aptos", size=10, color=self.COLORS["ink"])
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.border = Border(bottom=Side(style="hair", color=self.COLORS["line"]))
                if row_index % 2 == 0:
                    cell.fill = PatternFill("solid", fgColor=self.COLORS["paper"])

    def _draw_kpi(self, worksheet, cell_range, label, value, caption):
        start_cell, end_cell = cell_range.split(":")
        start = worksheet[start_cell]
        end = worksheet[end_cell]
        min_row, max_row = start.row, end.row
        min_column, max_column = start.column, end.column

        for row in worksheet.iter_rows(
            min_row=min_row,
            max_row=max_row,
            min_col=min_column,
            max_col=max_column,
        ):
            for cell in row:
                cell.fill = PatternFill("solid", fgColor=self.COLORS["white"])
                cell.border = self._thin_border()

        worksheet.merge_cells(
            start_row=min_row,
            start_column=min_column,
            end_row=min_row,
            end_column=max_column,
        )
        worksheet.merge_cells(
            start_row=min_row + 1,
            start_column=min_column,
            end_row=min_row + 2,
            end_column=max_column,
        )
        worksheet.merge_cells(
            start_row=max_row,
            start_column=min_column,
            end_row=max_row,
            end_column=max_column,
        )

        label_cell = worksheet.cell(min_row, min_column)
        label_cell.value = label
        label_cell.font = Font(name="Aptos", size=10, bold=True, color=self.COLORS["muted"])
        label_cell.alignment = Alignment(vertical="center")

        value_cell = worksheet.cell(min_row + 1, min_column)
        value_cell.value = value
        value_cell.font = Font(name="Aptos Display", size=22, bold=True, color=self.COLORS["teal"])
        value_cell.alignment = Alignment(vertical="center")

        caption_cell = worksheet.cell(max_row, min_column)
        caption_cell.value = caption
        caption_cell.font = Font(name="Aptos", size=9, color=self.COLORS["muted"])
        caption_cell.alignment = Alignment(vertical="center")

    def _write_chart_sources(self, worksheet, reports, summary):
        findings = [
            ("Duplicate PN", summary.get("Duplicate PN Groups", 0)),
            ("Near values", summary.get("Near Value Pairs", 0)),
            ("Package variants", summary.get("Different Package Groups", 0)),
            ("Voltage variants", summary.get("Different Voltage Groups", 0)),
            ("Material variants", summary.get("Different Material Groups", 0)),
        ]
        worksheet["J1"] = "Finding"
        worksheet["K1"] = "Count"
        for row_index, (label, count) in enumerate(findings, start=2):
            worksheet.cell(row_index, 10, label)
            worksheet.cell(row_index, 11, count)

        worksheet["M1"] = "Severity"
        worksheet["N1"] = "Count"
        rule_findings = reports.get("ai_rule_findings", pd.DataFrame())
        severity_counts = (
            rule_findings["Severity"].value_counts().to_dict()
            if "Severity" in rule_findings
            else {}
        )
        for row_index, severity in enumerate(("Critical", "High", "Medium", "Low"), start=2):
            worksheet.cell(row_index, 13, severity)
            worksheet.cell(row_index, 14, int(severity_counts.get(severity, 0)))

        worksheet["P1"] = "Vendor"
        worksheet["Q1"] = "Quantity"
        normalized = reports.get("normalized_bom", pd.DataFrame())
        if {"Vendor", "Quantity_Normalized"}.issubset(normalized.columns):
            vendor_quantities = (
                normalized[normalized["Vendor"].ne("Unknown")]
                .groupby("Vendor")["Quantity_Normalized"]
                .sum()
                .sort_values(ascending=False)
                .head(8)
            )
            for row_index, (vendor, quantity) in enumerate(vendor_quantities.items(), start=2):
                worksheet.cell(row_index, 16, self._safe_excel_value(vendor))
                worksheet.cell(row_index, 17, float(quantity))

    @staticmethod
    def _management_action_text(summary):
        risk_count = summary.get("High Risk Findings", 0)
        avl_count = summary.get("Unified AVL Ready", 0)
        merge_count = summary.get("Top Merge Candidates", 0)
        return (
            "MANAGEMENT ACTIONS\n\n"
            f"1. Resolve {risk_count} high-risk finding(s).\n\n"
            f"2. Approve {avl_count} unified AVL candidate(s).\n\n"
            f"3. Review {merge_count} top merge candidate(s).\n\n"
            "Engineering validation remains mandatory for protected circuits."
        )

    def _add_findings_chart(self, worksheet):
        chart = BarChart()
        chart.type = "col"
        chart.style = 10
        chart.title = "Normalization Findings"
        chart.y_axis.title = "Groups / pairs"
        chart.height = 7.2
        chart.width = 13.2
        chart.add_data(Reference(worksheet, min_col=11, min_row=1, max_row=6), titles_from_data=True)
        chart.set_categories(Reference(worksheet, min_col=10, min_row=2, max_row=6))
        chart.legend = None
        worksheet.add_chart(chart, "A14")

    def _add_risk_chart(self, worksheet):
        chart = PieChart()
        chart.style = 10
        chart.title = "Rule Findings by Severity"
        chart.height = 7.2
        chart.width = 10.2
        chart.add_data(Reference(worksheet, min_col=14, min_row=1, max_row=5), titles_from_data=True)
        chart.set_categories(Reference(worksheet, min_col=13, min_row=2, max_row=5))
        chart.dataLabels = DataLabelList()
        chart.dataLabels.showPercent = True
        chart.dataLabels.showLeaderLines = True
        worksheet.add_chart(chart, "E14")

    def _add_vendor_chart(self, worksheet):
        populated_rows = max(
            1,
            sum(1 for row in range(2, 10) if worksheet.cell(row, 16).value is not None),
        )
        chart = BarChart()
        chart.type = "bar"
        chart.style = 11
        chart.title = "Top Vendor Quantity Exposure"
        chart.x_axis.title = "Quantity"
        chart.height = 7.8
        chart.width = 17.5
        chart.add_data(
            Reference(worksheet, min_col=17, min_row=1, max_row=populated_rows + 1),
            titles_from_data=True,
        )
        chart.set_categories(
            Reference(worksheet, min_col=16, min_row=2, max_row=populated_rows + 1)
        )
        chart.legend = None
        worksheet.add_chart(chart, "A29")

    @staticmethod
    def _compact_spec(row):
        parts = [
            row.get("Normalized_Value", ""),
            row.get("Voltage", ""),
            row.get("Dielectric", ""),
            row.get("Tolerance", ""),
            row.get("Size", ""),
        ]
        text = " ".join(str(part).strip() for part in parts if str(part).strip())
        return text or row.get("Normalize_Key", "")

    @staticmethod
    def _split_spec(row):
        value = str(row.get("Normalized_Value", "") or "").strip()
        details = [
            row.get("Voltage", ""),
            row.get("Dielectric", ""),
            row.get("Tolerance", ""),
            row.get("Size", ""),
        ]
        spec_detail = " ".join(str(part).strip() for part in details if str(part).strip())
        return value or str(row.get("Normalize_Key", "") or ""), spec_detail

    @staticmethod
    def _compact_attribute_values(attribute_values):
        values = [value.strip() for value in str(attribute_values or "").split(",") if value.strip()]
        return "⇄".join(values) if values else ""

    @staticmethod
    def _compact_value_pair(values):
        cleaned_values = [str(value).strip() for value in values if str(value).strip()]
        return "⇄".join(cleaned_values[:3])

    @staticmethod
    def _display_resistance_value(value):
        text = str(value or "").strip()
        return text.replace("MOhm", "MΩ").replace("kOhm", "kΩ").replace("mOhm", "mΩ").replace("Ohm", "Ω")

    @staticmethod
    def _nearby_family(row):
        fields = [
            row.get("Tolerance", ""),
            row.get("Power_Rating", ""),
            row.get("Material", ""),
            row.get("Size", ""),
        ]
        return " ".join(str(value).strip() for value in fields if str(value).strip())

    @staticmethod
    def _signed_percent(value):
        try:
            numeric_value = float(value)
        except (TypeError, ValueError):
            return ""
        rounded = round(numeric_value, 2)
        text = f"{rounded:+.2f}" if rounded > 0 else f"{rounded:.2f}"
        text = text.rstrip("0").rstrip(".")
        return f"{text}%"

    @staticmethod
    def _quantity_share(quantity, total_quantity):
        if not total_quantity:
            return "0 (0%)"
        percent = round(float(quantity) / float(total_quantity) * 100)
        return f"{float(quantity):g} ({percent:g}%)"

    def _target_candidate_share(self, group):
        target_quantity = group["target"]["quantity"]
        candidate_quantity = sum(candidate["quantity"] for candidate in group["candidates"])
        total_quantity = group["total_quantity"]
        return (
            self._quantity_share(target_quantity, total_quantity),
            self._quantity_share(candidate_quantity, total_quantity),
        )

    @staticmethod
    def _difficulty_label(spec_similarity):
        try:
            score = float(spec_similarity)
        except (TypeError, ValueError):
            score = 0
        if score >= 95:
            return "★★★★★ Easy"
        if score >= 70:
            return "★★★★ Medium"
        if score >= 40:
            return "★★★ Hard"
        return "★ High risk"

    @staticmethod
    def _difference_label(spec_similarity):
        try:
            score = float(spec_similarity)
        except (TypeError, ValueError):
            score = 0
        if score == 100:
            return "🟢 Same spec"
        if score == 95:
            return "🟢 Vendor"
        if score == 70:
            return "🟠 Voltage"
        if score == 60:
            return "🟡 Tolerance"
        if score == 40:
            return "🟡 Package"
        return "🔴 Material / Value"

    @staticmethod
    def _reference_count(reference_text):
        if not isinstance(reference_text, str) or not reference_text.strip():
            return 0
        return len([reference for reference in re.split(r"[,;\s]+", reference_text.strip()) if reference])

    @staticmethod
    def _unique_text(series, case_insensitive=False):
        values = []
        seen = set()
        for value in series:
            text = "" if pd.isna(value) else str(value).strip()
            identity = text.casefold() if case_insensitive else text
            if text and identity not in seen:
                values.append(text)
                seen.add(identity)
        return values

    @staticmethod
    def _project_name(reports):
        metadata = ExcelReportWriter._metric_lookup(
            reports.get("report_metadata", pd.DataFrame()),
            key_column="Property",
        )
        project_name = str(metadata.get("Project Name", "")).strip()
        if project_name:
            return project_name
        input_file = str(metadata.get("Input File", "")).strip()
        if input_file:
            return Path(input_file).stem
        return "BOM Project"

    @staticmethod
    def _component_part_count(normalized, component_type):
        required_columns = {"Component_Type", "Part_Number"}
        if normalized.empty or not required_columns.issubset(normalized.columns):
            return 0
        part_numbers = normalized.loc[
            normalized["Component_Type"].eq(component_type),
            "Part_Number",
        ].replace("", pd.NA)
        return int(part_numbers.nunique())

    def _link_summary_to_detail(self, workbook):
        if "Capacitor Summary" not in workbook or "Merge Workspace" not in workbook:
            return

        summary = workbook["Capacitor Summary"]
        detail = workbook["Merge Workspace"]
        detail_rows = {
            detail.cell(row_index, 1).value: row_index
            for row_index in range(2, detail.max_row + 1)
            if detail.cell(row_index, 1).value
        }
        headers = [cell.value for cell in summary[1]]
        try:
            detail_column = headers.index("Detail") + 1
            group_column = headers.index("Merge ID") + 1
        except ValueError:
            return

        for row_index in range(2, summary.max_row + 1):
            group_id = summary.cell(row_index, group_column).value
            target_row = detail_rows.get(group_id, 1)
            cell = summary.cell(row_index, detail_column)
            cell.hyperlink = f"#'Merge Workspace'!A{target_row}"
            cell.style = "Hyperlink"

    def _sync_summary_from_workspace(self, workbook):
        """Make Summary decision read from Merge Workspace by Merge ID."""
        if "Capacitor Summary" not in workbook or "Merge Workspace" not in workbook:
            return
        summary = workbook["Capacitor Summary"]
        summary_headers = [cell.value for cell in summary[1]]
        try:
            merge_id_column = summary_headers.index("Merge ID") + 1
            decision_column = summary_headers.index("RD Decision") + 1
        except ValueError:
            return
        for row_index in range(2, summary.max_row + 1):
            merge_id_cell = summary.cell(row_index, merge_id_column).coordinate
            summary.cell(row_index, decision_column).value = f'=IFERROR(IF(VLOOKUP(${merge_id_cell},\'Merge Workspace\'!$A:$K,11,FALSE)="","Pending",VLOOKUP(${merge_id_cell},\'Merge Workspace\'!$A:$K,11,FALSE)),"Pending")'

    def _link_resistor_summary_to_detail(self, workbook):
        if "Resistor Summary" not in workbook or "Resistor Detail" not in workbook:
            return

        summary = workbook["Resistor Summary"]
        detail = workbook["Resistor Detail"]
        detail_rows = {
            detail.cell(row_index, 1).value: row_index
            for row_index in range(2, detail.max_row + 1)
            if detail.cell(row_index, 2).value == "Value Header"
        }
        headers = [cell.value for cell in summary[1]]
        try:
            detail_column = headers.index("Detail") + 1
            group_column = headers.index("Group") + 1
        except ValueError:
            return

        for row_index in range(2, summary.max_row + 1):
            group_id = summary.cell(row_index, group_column).value
            target_row = detail_rows.get(group_id, 1)
            cell = summary.cell(row_index, detail_column)
            cell.hyperlink = f"#'Resistor Detail'!A{target_row}"
            cell.style = "Hyperlink"

    def _add_rd_decision_dropdowns(self, workbook):
        decision_options = "Merge,Keep,Skip"
        for sheet_name in ("Merge Workspace",):
            if sheet_name not in workbook:
                continue
            worksheet = workbook[sheet_name]
            headers = [cell.value for cell in worksheet[1]]
            if "RD Decision" not in headers:
                continue
            column_letter = worksheet.cell(1, headers.index("RD Decision") + 1).column_letter
            validation = DataValidation(type="list", formula1=f'"{decision_options}"', allow_blank=True)
            validation.error = "Please choose Merge, Keep, or Skip."
            validation.errorTitle = "Invalid RD Decision"
            validation.prompt = "Select the engineering decision for this merge row."
            validation.promptTitle = "RD Decision"
            worksheet.add_data_validation(validation)
            validation.add(f"{column_letter}2:{column_letter}1048576")

    def _format_data_sheet(self, worksheet, dataframe, report_key):
        worksheet.sheet_view.showGridLines = False
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions
        worksheet.row_dimensions[1].height = 30

        header_fill = PatternFill("solid", fgColor=self.COLORS["navy"])
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = Font(name="Aptos", size=10, bold=True, color=self.COLORS["white"])
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for column_index, column_name in enumerate(dataframe.columns, start=1):
            values = [str(column_name)]
            values.extend(str(value) for value in dataframe[column_name].head(250).fillna(""))
            width = min(max(max(len(value) for value in values) + 2, 11), 48)
            worksheet.column_dimensions[worksheet.cell(1, column_index).column_letter].width = width

        for row_index, row in enumerate(worksheet.iter_rows(min_row=2), start=2):
            if row_index % 2 == 0:
                for cell in row:
                    cell.fill = PatternFill("solid", fgColor=self.COLORS["paper"])
            for cell in row:
                cell.font = Font(name="Aptos", size=10, color=self.COLORS["ink"])
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.border = Border(bottom=Side(style="hair", color=self.COLORS["line"]))

        self._add_conditional_formatting(worksheet, dataframe)
        if report_key == "specification_summary":
            self._format_specification_summary(worksheet, dataframe)
        if report_key == "specification_detail":
            self._format_specification_detail(worksheet, dataframe)
        if report_key == "resistor_summary":
            self._format_resistor_summary(worksheet, dataframe)
        if report_key == "resistor_detail":
            self._format_resistor_detail(worksheet, dataframe)
        if report_key == "nearby_value":
            self._format_nearby_value(worksheet, dataframe)
        if report_key in {"summary", "bom_score", "report_metadata"}:
            for cell in worksheet["A"]:
                cell.font = Font(name="Aptos", size=10, bold=True, color=self.COLORS["ink"])

    def _format_specification_summary(self, worksheet, dataframe):
        headers = [cell.value for cell in worksheet[1]]
        if dataframe.empty:
            return
        widths = {
            "Merge ID": 12,
            "Review Item": 28,
            "BOM Qty": 10,
            "Keep Qty": 10,
            "Merge Qty": 10,
            "Priority": 12,
            "Why Review": 32,
            "RD Decision": 14,
            "Detail": 10,
        }
        for header, width in widths.items():
            if header in headers:
                worksheet.column_dimensions[worksheet.cell(1, headers.index(header) + 1).column_letter].width = width
        if "Review Item" in headers:
            value_letter = worksheet.cell(1, headers.index("Review Item") + 1).column_letter
            worksheet.column_dimensions[value_letter].width = 16
            for cell in worksheet[value_letter][1:]:
                cell.font = Font(name="Aptos Display", size=12, bold=True, color=self.COLORS["teal"])

    def _format_specification_detail(self, worksheet, dataframe):
        if dataframe.empty:
            return
        headers = [cell.value for cell in worksheet[1]]
        try:
            merge_id_column = headers.index("Merge ID") + 1
            merge_pn_column = headers.index("Merge PN") + 1
        except ValueError:
            return

        fills = {
            "KEEP": PatternFill("solid", fgColor=self.COLORS["low"]),
            "MERGE": PatternFill("solid", fgColor=self.COLORS["medium"]),
        }
        for row_index in range(2, worksheet.max_row + 1):
            section = "MERGE" if worksheet.cell(row_index, merge_pn_column).value else "KEEP"
            fill = fills.get(section)
            if fill:
                for cell in worksheet[row_index]:
                    cell.fill = fill
            if section == "KEEP":
                worksheet.cell(row_index, merge_id_column).font = Font(
                    name="Aptos",
                    size=10,
                    bold=True,
                    color=self.COLORS["ink"],
                )

        widths = {
            "Merge ID": 12,
            "Keep PN": 18,
            "Merge PN": 18,
            "Keep Qty": 10,
            "Merge Qty": 10,
            "Difference": 28,
            "Vendor": 18,
            "Package": 18,
            "Voltage": 16,
            "Material": 16,
            "RD Decision": 14,
        }
        for header, width in widths.items():
            if header in headers:
                worksheet.column_dimensions[worksheet.cell(1, headers.index(header) + 1).column_letter].width = width

    def _format_resistor_summary(self, worksheet, dataframe):
        headers = [cell.value for cell in worksheet[1]]
        if "Group" in headers:
            group_letter = worksheet.cell(1, headers.index("Group") + 1).column_letter
            worksheet.column_dimensions[group_letter].hidden = True
        if dataframe.empty:
            return
        widths = {
            "Value": 14,
            "PN Count": 10,
            "Total Qty": 10,
            "Action / Target PN": 18,
            "Priority": 12,
            "Reason (相似度分类)": 32,
            "Detail": 12,
            "RD Decision": 22,
        }
        for header, width in widths.items():
            if header in headers:
                worksheet.column_dimensions[worksheet.cell(1, headers.index(header) + 1).column_letter].width = width
        if "Value" in headers:
            value_letter = worksheet.cell(1, headers.index("Value") + 1).column_letter
            for cell in worksheet[value_letter][1:]:
                cell.font = Font(name="Aptos Display", size=12, bold=True, color=self.COLORS["teal"])
        if "Action / Target PN" in headers:
            action_column = headers.index("Action / Target PN") + 1
            fills = {
                "Review": PatternFill("solid", fgColor=self.COLORS["medium"]),
                "Review Required": PatternFill("solid", fgColor=self.COLORS["medium"]),
            }
            for row_index in range(2, worksheet.max_row + 1):
                fill = fills.get(worksheet.cell(row_index, action_column).value)
                if fill:
                    worksheet.cell(row_index, action_column).fill = fill

    def _format_resistor_detail(self, worksheet, dataframe):
        if dataframe.empty:
            return
        headers = [cell.value for cell in worksheet[1]]
        if "Group" in headers:
            group_letter = worksheet.cell(1, headers.index("Group") + 1).column_letter
            worksheet.column_dimensions[group_letter].hidden = True
        widths = {
            "Value": 14,
            "Spec": 34,
            "PN": 18,
            "Qty": 10,
            "Status": 14,
            "Difference": 18,
            "Why Listed": 18,
        }
        for header, width in widths.items():
            if header in headers:
                worksheet.column_dimensions[worksheet.cell(1, headers.index(header) + 1).column_letter].width = width
        try:
            row_type_column = headers.index("Row Type") + 1
        except ValueError:
            return
        fills = {
            "Value Header": PatternFill("solid", fgColor=self.COLORS["cyan"]),
            "Spec Header": PatternFill("solid", fgColor="E7E9EC"),
            "PN": PatternFill("solid", fgColor=self.COLORS["white"]),
        }
        for row_index in range(2, worksheet.max_row + 1):
            row_type = worksheet.cell(row_index, row_type_column).value
            fill = fills.get(row_type)
            if fill:
                for cell in worksheet[row_index]:
                    cell.fill = fill
            if row_type in {"Value Header", "Spec Header"}:
                for cell in worksheet[row_index]:
                    cell.font = Font(name="Aptos", size=10, bold=True, color=self.COLORS["ink"])

    def _format_nearby_value(self, worksheet, dataframe):
        if dataframe.empty:
            return
        widths = {
            "Current Value": 16,
            "Current BOM Qty": 14,
            "Nearby Value": 16,
            "Candidate Qty": 14,
            "Difference": 12,
            "Tolerance Band": 14,
            "Family": 34,
            "Candidate PNs": 32,
        }
        headers = [cell.value for cell in worksheet[1]]
        for header, width in widths.items():
            if header in headers:
                worksheet.column_dimensions[worksheet.cell(1, headers.index(header) + 1).column_letter].width = width

    def _add_conditional_formatting(self, worksheet, dataframe):
        if dataframe.empty:
            return

        column_lookup = {
            str(cell.value): cell.column_letter
            for cell in worksheet[1]
            if cell.value is not None
        }
        last_row = worksheet.max_row

        severity_column = column_lookup.get("Severity")
        if severity_column:
            fills = {
                "Critical": self.COLORS["critical"],
                "High": self.COLORS["high"],
                "Medium": self.COLORS["medium"],
                "Low": self.COLORS["low"],
            }
            for severity, color in fills.items():
                worksheet.conditional_formatting.add(
                    f"{severity_column}2:{severity_column}{last_row}",
                    FormulaRule(
                        formula=[f'${severity_column}2="{severity}"'],
                        fill=PatternFill("solid", fgColor=color),
                    ),
                )

        for column_name in ("Opportunity_Score", "Data_Quality_Score", "Confidence"):
            column_letter = column_lookup.get(column_name)
            if column_letter:
                worksheet.conditional_formatting.add(
                    f"{column_letter}2:{column_letter}{last_row}",
                    ColorScaleRule(
                        start_type="min",
                        start_color="F8696B",
                        mid_type="percentile",
                        mid_value=50,
                        mid_color="FFEB84",
                        end_type="max",
                        end_color="63BE7B",
                    ),
                )

        critical_column = column_lookup.get("Has_Critical_Part")
        if critical_column:
            worksheet.conditional_formatting.add(
                f"{critical_column}2:{critical_column}{last_row}",
                FormulaRule(
                    formula=[f'${critical_column}2="Yes"'],
                    fill=PatternFill("solid", fgColor=self.COLORS["critical"]),
                ),
            )

    def _sanitize_dataframe(self, dataframe):
        safe_dataframe = dataframe.copy()
        safe_dataframe.columns = self._safe_column_labels(safe_dataframe.columns)
        for column in safe_dataframe.columns:
            safe_dataframe[column] = safe_dataframe[column].map(self._safe_excel_value)
        return safe_dataframe

    def _safe_column_labels(self, columns):
        counts = {}
        labels = []
        for index, column in enumerate(columns, start=1):
            label = str(self._safe_excel_value(str(column))).strip()
            if not label:
                label = f"Column_{index}"
            counts[label] = counts.get(label, 0) + 1
            labels.append(label if counts[label] == 1 else f"{label}_{counts[label]}")
        return labels

    @staticmethod
    def _safe_excel_value(value):
        if not isinstance(value, str):
            return value
        formula_candidate = value.lstrip("\t\r\n\ufeff")
        if formula_candidate.startswith(("=", "+", "-", "@")):
            return f"'{value}"
        return value

    @staticmethod
    def _metric_lookup(dataframe, key_column="Metric"):
        if dataframe.empty or key_column not in dataframe or "Value" not in dataframe:
            return {}
        return dict(zip(dataframe[key_column], dataframe["Value"]))

    def _thin_border(self):
        side = Side(style="thin", color=self.COLORS["line"])
        return Border(left=side, right=side, top=side, bottom=side)