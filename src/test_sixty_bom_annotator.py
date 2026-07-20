import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook

from sixty_bom_annotator import SixtyBOMAnnotator


class SixtyBOMAnnotatorTests(unittest.TestCase):
    def setUp(self):
        self.headers = [
            "Top Assembly Item",
            "BOM Level",
            "Seq Num",
            "Comp Type",
            "Substitute Priority",
            "Item Number",
            "Item Description",
            "Comp Quantity",
            "插件位置",
            "OpenFlag",
            "備註",
            "料件狀態",
            "承認狀態",
            "燒錄狀態",
            "BOM Owner",
        ]

    def test_annotates_master_substitute_and_near_value_rows(self):
        with tempfile.TemporaryDirectory() as temporary_directory:
            input_file = Path(temporary_directory) / "60Bom.xlsx"
            output_file = Path(temporary_directory) / "60Bom_歸一化標註.xlsx"
            self._create_workbook(input_file)

            annotator = SixtyBOMAnnotator()
            self.assertTrue(annotator.supports_workbook(input_file))
            summary = annotator.annotate(input_file, output_file)

            workbook = load_workbook(output_file, read_only=True, data_only=True)
            worksheet = workbook["My Excel Sheet"]
            headers = [cell.value for cell in worksheet[1]]
            rows = list(worksheet.iter_rows(min_row=2, values_only=True))
            by_part_number = {row[5]: row for row in rows}
            judgment_column = headers.index("歸一化判定")
            value_column = headers.index("歸一化值")
            near_value_column = headers.index("相近值提示")
            group_column = headers.index("歸一化群組(Group ID)")
            members_column = headers.index("群組成員(主料)")
            risk_category_column = headers.index("風險分類")
            risk_reason_column = headers.index("風險原因")
            decision_column = headers.index("判定燈號")
            check_report_title = workbook["檢查報告"]["A1"].value
            workbook.close()

            self.assertEqual(headers[-len(annotator.ANALYSIS_COLUMNS):], list(annotator.ANALYSIS_COLUMNS))
            self.assertEqual(by_part_number["R-MAIN"][value_column], "73.2千歐姆")
            self.assertEqual(by_part_number["R-MAIN"][judgment_column], "可評估歸一化")
            self.assertEqual(by_part_number["R-MAIN2"][judgment_column], "可評估歸一化")
            self.assertEqual(by_part_number["R-MAIN"][group_column], "RC0001")
            self.assertEqual(by_part_number["R-MAIN2"][group_column], "RC0001")
            self.assertIn("R-MAIN", by_part_number["R-MAIN"][members_column])
            self.assertIn("R-MAIN2", by_part_number["R-MAIN2"][members_column])
            self.assertEqual(by_part_number["R-MAIN"][risk_category_column], "Vendor")
            self.assertIn("供應商策略", by_part_number["R-MAIN"][risk_reason_column])
            self.assertEqual(by_part_number["R-MAIN"][decision_column], "🟢 可以統一")
            self.assertEqual(by_part_number["C-ALT"][value_column], "100納法")
            self.assertEqual(by_part_number["C-ALT"][judgment_column], "Second Source（不納入比較）")
            self.assertEqual(by_part_number["M-ALT"][judgment_column], "Second Source（不納入比較）")
            self.assertEqual(by_part_number["R-10K"][judgment_column], "相近值需確認")
            self.assertIn("標準系列：E24", by_part_number["R-10K"][near_value_column])
            self.assertIn("附近精細值(E96)", by_part_number["R-10K"][near_value_column])
            self.assertIn("12K", by_part_number["R-10K"][near_value_column])
            self.assertEqual(check_report_title, "Original check report")
            self.assertEqual(summary["electronic_rows"], 7)
            self.assertEqual(summary["substitute_rows"], 3)
            self.assertEqual(summary["normalization_groups"], 1)

    def test_preserves_lowercase_milliohm_notation(self):
        raw_value, normalized_value, numeric_value = SixtyBOMAnnotator._extract_resistance(
            "RES 1m OHM 3W (2512) 1%"
        )

        self.assertEqual(raw_value, "1m OHM")
        self.assertEqual(normalized_value, "1mOhm")
        self.assertEqual(numeric_value, 0.001)

    def test_second_source_is_not_used_for_main_part_comparison(self):
        with tempfile.TemporaryDirectory() as temporary_directory:
            input_file = Path(temporary_directory) / "60Bom.xlsx"
            output_file = Path(temporary_directory) / "60Bom_歸一化標註.xlsx"

            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "My Excel Sheet"
            worksheet.append(self.headers)
            worksheet.append(self._row(10, "", "R-MAIN", "RES 10K OHM 1/16W (0402) 1%", "R1"))
            worksheet.append(self._row(10, "S", "R-S", "RES 10K OHM 1/16W (0402) 1%", ""))
            workbook.save(input_file)
            workbook.close()

            SixtyBOMAnnotator().annotate(input_file, output_file)
            workbook = load_workbook(output_file, read_only=True, data_only=True)
            worksheet = workbook["My Excel Sheet"]
            headers = [cell.value for cell in worksheet[1]]
            rows = list(worksheet.iter_rows(min_row=2, values_only=True))
            by_part_number = {row[5]: row for row in rows}
            judgment_column = headers.index("歸一化判定")
            workbook.close()

            self.assertEqual(by_part_number["R-MAIN"][judgment_column], "規格已標準化")
            self.assertEqual(by_part_number["R-S"][judgment_column], "Second Source（不納入比較）")

    def _create_workbook(self, workbook_file):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "My Excel Sheet"
        worksheet.append(self.headers)
        worksheet.append(self._row(10, "", "R-MAIN", "RES 73.2K OHM 1/16W (0402) 1%//RALEC/RTT", "R1"))
        worksheet.append(self._row(10, "S", "R-ALT", "RES 73.2K OHM 1/16W (0402) 1%//TA-I/RM", ""))
        worksheet.append(self._row(20, "", "C-MAIN", "MLCC 0.1UF/16V (0201) X5R 10%//MURATA/GRM", "C1"))
        worksheet.append(self._row(20, "S", "C-ALT", "MLCC 100NF/16V (0201) X5R 10%//SAMSUNG/CL", ""))
        worksheet.append(self._row(30, "", "R-10K", "RES 10K OHM 1/16W (0402) 1%//RALEC/RTT", "R2"))
        worksheet.append(self._row(40, "", "R-12K", "RES 12K OHM 1/16W (0402) 1%//RALEC/RTT", "R3"))
        worksheet.append(self._row(50, "", "R-MAIN2", "RES 73.2K OHM 1/16W (0402) 1%//UNI-OHM/0402", "R4"))
        worksheet.append(self._row(60, "S", "M-ALT", "MECHANICAL PART//VENDOR", ""))
        report = workbook.create_sheet("檢查報告")
        report["A1"] = "Original check report"
        workbook.save(workbook_file)
        workbook.close()

    @staticmethod
    def _row(sequence, component_type, part_number, description, plugin_location):
        return [
            "P500MV",
            2,
            sequence,
            component_type,
            100 if component_type else None,
            part_number,
            description,
            1,
            plugin_location,
            None,
            None,
            "Active",
            "Y",
            "N",
            "RD",
        ]


if __name__ == "__main__":
    unittest.main()