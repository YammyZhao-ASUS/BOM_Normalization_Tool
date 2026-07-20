import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook

from bom_intelligence import BOMIntelligencePlatform
from bom_reader import BOMReadError, BOMReader


class BOMReaderTests(unittest.TestCase):
    def test_detects_bom_sheet_and_shifted_header(self):
        with tempfile.TemporaryDirectory() as temporary_directory:
            input_file = Path(temporary_directory) / "shifted.xlsx"
            workbook = Workbook()
            workbook.active.title = "Cover"
            workbook.active["A1"] = "Project cover"
            worksheet = workbook.create_sheet("Production BOM")
            worksheet.append(["Project", "Board A"])
            worksheet.append([])
            worksheet.append(
                ["Part Number", "Part Reference", "Component_Name", "Quantity"]
            )
            worksheet.append(["PN1", "C1", "MLCC 100NF 16V 0402 X7R", 2])
            workbook.save(input_file)
            workbook.close()

            reports = BOMIntelligencePlatform().analyze_file(input_file)
            metadata = dict(
                zip(
                    reports["report_metadata"]["Property"],
                    reports["report_metadata"]["Value"],
                )
            )

            self.assertEqual(len(reports["normalized_bom"]), 1)
            self.assertEqual(metadata["Selected Sheet"], "Production BOM")
            self.assertEqual(metadata["Detected Header Row"], 3)

            renamed_file = input_file.with_name("renamed.xlsx")
            input_file.rename(renamed_file)
            self.assertTrue(renamed_file.is_file())

    def test_reads_shifted_csv_header(self):
        with tempfile.TemporaryDirectory() as temporary_directory:
            input_file = Path(temporary_directory) / "bom.csv"
            input_file.write_text(
                "Project,Board A,,\n"
                "Part Number,Part Reference,Component_Name,Quantity\n"
                "PR1,R1,RES 10K 0402,3\n",
                encoding="utf-8",
            )

            reader = BOMReader(
                input_file,
                BOMIntelligencePlatform.COLUMN_CANDIDATES,
            )
            dataframe = reader.load()

            self.assertEqual(len(dataframe), 1)
            self.assertEqual(dataframe.iloc[0]["Part Number"], "PR1")
            self.assertEqual(reader.header_row, 1)

    def test_rejects_workbook_without_recognizable_bom_header(self):
        with tempfile.TemporaryDirectory() as temporary_directory:
            input_file = Path(temporary_directory) / "not_a_bom.xlsx"
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.append(["Project summary", "Owner"])
            worksheet.append(["Board A", "Engineering"])
            workbook.save(input_file)
            workbook.close()

            with self.assertRaisesRegex(BOMReadError, "At least two recognized BOM fields"):
                BOMReader(
                    input_file,
                    BOMIntelligencePlatform.COLUMN_CANDIDATES,
                ).load()


if __name__ == "__main__":
    unittest.main()