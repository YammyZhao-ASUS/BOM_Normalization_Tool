import itertools
import math
import re
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

try:
    from .capacitor_normalizer import CapacitorNormalizer
except ImportError:
    from capacitor_normalizer import CapacitorNormalizer


class SixtyBOMAnnotator:
    """Append normalization review judgments to an ASUS 60BOM workbook."""

    DATA_SHEET_NAME = "My Excel Sheet"
    REQUIRED_COLUMNS = {
        "BOM Level",
        "Seq Num",
        "Comp Type",
        "Item Number",
        "Item Description",
        "Comp Quantity",
        "插件位置",
    }
    ANALYSIS_COLUMNS = (
        "元件類型",
        "60BOM料件角色",
        "原始規格值",
        "歸一化值",
        "額定電壓",
        "介質",
        "封裝/尺寸",
        "容差",
        "功率",
        "歸一化規格鍵",
        "歸一化群組(Group ID)",
        "群組成員(主料)",
        "相近值提示",
        "歸一化判定",
        "歸一化建議",
        "風險分類",
        "風險原因",
        "判定燈號",
    )
    DECISION_COLORS = {
        "green": "D9EAD3",
        "yellow": "FFF2CC",
        "red": "F4CCCC",
        "gray": "E7E6E6",
    }
    E12_VALUES = (1.0, 1.2, 1.5, 1.8, 2.2, 2.7, 3.3, 3.9, 4.7, 5.6, 6.8, 8.2)
    _E_SERIES_CACHE = {}
    CRITICAL_KEYWORDS = (
        "CPU",
        "VCORE",
        "DDR",
        "PCIE",
        "PCI-E",
        "USB",
        "TYPE-C",
        "TYPEC",
        "HDMI",
        "DISPLAYPORT",
        "WLAN",
        "WI-FI",
        "WIFI",
        "RF",
        "ANTENNA",
    )

    def __init__(self, near_value_ratio=2.2):
        if near_value_ratio <= 1:
            raise ValueError("near_value_ratio must be greater than 1")

        self.near_value_ratio = near_value_ratio
        self.capacitor_normalizer = CapacitorNormalizer()
        self._decision_fills = {
            level: PatternFill("solid", fgColor=color)
            for level, color in self.DECISION_COLORS.items()
        }

    @classmethod
    def supports_workbook(cls, input_file):
        workbook = load_workbook(input_file, read_only=True, data_only=False)
        try:
            if cls.DATA_SHEET_NAME not in workbook.sheetnames:
                return False

            worksheet = workbook[cls.DATA_SHEET_NAME]
            headers = {
                str(cell.value).strip()
                for cell in worksheet[1]
                if cell.value is not None
            }
            return cls.REQUIRED_COLUMNS.issubset(headers)
        finally:
            workbook.close()

    def annotate(self, input_file, output_file):
        input_path = Path(input_file)
        output_path = Path(output_file)
        workbook = load_workbook(input_path)

        try:
            worksheet = self._get_data_worksheet(workbook)
            column_map = self._header_map(worksheet)
            records = self._read_records(worksheet, column_map)
            self._link_substitutes(records)
            self._find_same_specification_candidates(records)
            self._assign_normalization_groups(records)
            self._find_near_value_candidates(records)
            self._set_judgments(records)
            self._write_annotations(worksheet, records)

            output_path.parent.mkdir(parents=True, exist_ok=True)
            workbook.save(output_path)
            return self._summary(records, output_path)
        finally:
            workbook.close()

    def _get_data_worksheet(self, workbook):
        if self.DATA_SHEET_NAME not in workbook.sheetnames:
            raise ValueError(f"60BOM data worksheet not found: {self.DATA_SHEET_NAME}")

        worksheet = workbook[self.DATA_SHEET_NAME]
        headers = set(self._header_map(worksheet))
        missing_columns = self.REQUIRED_COLUMNS - headers
        if missing_columns:
            missing = ", ".join(sorted(missing_columns))
            raise ValueError(f"60BOM required columns are missing: {missing}")

        return worksheet

    @staticmethod
    def _header_map(worksheet):
        return {
            str(cell.value).strip(): cell.column
            for cell in worksheet[1]
            if cell.value is not None and str(cell.value).strip()
        }

    def _read_records(self, worksheet, column_map):
        records = []

        for row_number in range(2, worksheet.max_row + 1):
            item_number = self._cell_text(worksheet, row_number, column_map["Item Number"])
            description = self._cell_text(worksheet, row_number, column_map["Item Description"])
            if not item_number and not description:
                continue

            comp_type = self._cell_text(worksheet, row_number, column_map["Comp Type"])
            role = "Second Source" if comp_type.upper() == "S" else "主料"
            spec_text = description.split("//", 1)[0].strip()
            component_type = self._detect_component_type(spec_text, self._cell_text(worksheet, row_number, column_map["插件位置"]))
            details = self._extract_details(spec_text, description, component_type)
            record = {
                "row": row_number,
                "item_number": item_number,
                "description": description,
                "spec_text": spec_text,
                "bom_level": self._cell_text(worksheet, row_number, column_map["BOM Level"]),
                "seq_num": self._cell_text(worksheet, row_number, column_map["Seq Num"]),
                "plugin_location": self._cell_text(worksheet, row_number, column_map["插件位置"]),
                "role": role,
                "component_type": component_type,
                "group_id": "",
                "substitute_count": 0,
                "substitute_differences": [],
                "same_specification_groups": 0,
                "normalize_group_id": "",
                "normalize_group_members": [],
                "near_values": [],
                "is_critical": self._is_critical(spec_text),
                "risk_category": "",
                "risk_reason": "",
                "decision_label": "",
                "decision_color": "yellow",
            }
            record.update(details)
            records.append(record)

        return records

    @staticmethod
    def _cell_text(worksheet, row_number, column_number):
        value = worksheet.cell(row_number, column_number).value
        if value is None:
            return ""

        return re.sub(r"\s+", " ", str(value)).strip()

    @staticmethod
    def _detect_component_type(spec_text, plugin_location):
        upper_spec = spec_text.upper()
        upper_location = plugin_location.upper()

        if re.search(r"\b(?:MLCC|CAP(?:ACITOR)?|ECAP|TANT)\b", upper_spec) or re.search(
            r"(?:^|[,;/\s])(?:C|CE)\d+", upper_location
        ):
            return "電容"

        if re.search(r"\bRES(?:ISTOR)?\b", upper_spec) or re.search(
            r"(?:^|[,;/\s])(?:R|PR)\d+", upper_location
        ):
            return "電阻"

        return ""

    def _extract_details(self, spec_text, full_description, component_type):
        if component_type == "電容":
            return self._extract_capacitor_details(spec_text, full_description)

        if component_type == "電阻":
            return self._extract_resistor_details(spec_text, full_description)

        return {
            "original_value": "",
            "normalized_value": "",
            "numeric_value": None,
            "voltage": "",
            "dielectric": "",
            "package": "",
            "tolerance": "",
            "power": "",
            "technology": "",
            "normalize_key": "",
            "missing_fields": [],
        }

    def _extract_capacitor_details(self, spec_text, full_description):
        raw_value = self._extract_capacitor_value(spec_text)
        normalized_value = self.capacitor_normalizer.normalize(raw_value) if raw_value else ""
        voltage = self._extract_voltage(spec_text)
        dielectric = self._extract_dielectric(spec_text)
        package = self._extract_package(spec_text)
        tolerance = self._extract_tolerance(full_description)
        technology = self._extract_capacitor_technology(spec_text)
        missing_fields = [field for field, value in (
            ("電容量", normalized_value),
            ("額定電壓", voltage),
            ("封裝/尺寸", package),
            ("容差", tolerance),
        ) if not value]
        if technology == "MLCC" and not dielectric:
            missing_fields.append("介質")

        normalize_key = "|".join((
            "C",
            technology or "-",
            normalized_value or "-",
            voltage or "-",
            dielectric or "-",
            package or "-",
            tolerance or "-",
        ))
        return {
            "original_value": raw_value,
            "normalized_value": normalized_value,
            "numeric_value": self._capacitance_to_pf(normalized_value),
            "voltage": voltage,
            "dielectric": dielectric,
            "package": package,
            "tolerance": tolerance,
            "power": "",
            "technology": technology,
            "normalize_key": normalize_key,
            "missing_fields": missing_fields,
        }

    def _extract_resistor_details(self, spec_text, full_description):
        raw_value, normalized_value, numeric_value = self._extract_resistance(spec_text)
        package = self._extract_package(spec_text)
        tolerance = self._extract_tolerance(full_description)
        power = self._extract_power(spec_text)
        missing_fields = [field for field, value in (
            ("阻值", normalized_value),
            ("封裝/尺寸", package),
            ("容差", tolerance),
            ("功率", power),
        ) if not value]
        normalize_key = "|".join((
            "R",
            normalized_value or "-",
            tolerance or "-",
            power or "-",
            package or "-",
        ))
        return {
            "original_value": raw_value,
            "normalized_value": normalized_value,
            "numeric_value": numeric_value,
            "voltage": "",
            "dielectric": "",
            "package": package,
            "tolerance": tolerance,
            "power": power,
            "technology": "",
            "normalize_key": normalize_key,
            "missing_fields": missing_fields,
        }

    @staticmethod
    def _extract_capacitor_value(spec_text):
        normalized_text = spec_text.upper().replace("μ", "U").replace("µ", "U")
        match = re.search(r"\b(\d+(?:\.\d+)?)\s*(UF|NF|PF|P)(?=\s|/|\(|$)", normalized_text)
        if match:
            return "".join(match.groups())

        match = re.search(r"\b(\d{3})\b", normalized_text)
        return match.group(1) if match else ""

    @staticmethod
    def _extract_voltage(spec_text):
        match = re.search(r"\b(\d+(?:\.\d+)?)\s*V\b", spec_text.upper())
        return f"{match.group(1)}V" if match else ""

    @staticmethod
    def _extract_dielectric(spec_text):
        upper_spec = spec_text.upper()
        for value in ("X5R", "X6S", "X7R", "NP0", "NPO", "C0G", "COG"):
            if value in upper_spec:
                return "NP0" if value in {"NPO", "C0G", "COG"} else value

        return ""

    @staticmethod
    def _extract_package(spec_text):
        standard_size = re.search(r"\b(0201|0402|0603|0805|1206|1210|1812|2220|2512)\b", spec_text)
        if standard_size:
            return standard_size.group(1)

        case_size = re.search(r"\b(\d{4}(?:/[A-Z])?)\b", spec_text.upper())
        if case_size and case_size.group(1) not in {"0201", "0402", "0603", "0805", "1206", "1210", "1812", "2220", "2512"}:
            return case_size.group(1)

        physical_size = re.search(r"\b(\d+(?:\.\d+)?)\s*[*X]\s*(\d+(?:\.\d+)?)\b", spec_text.upper())
        if physical_size:
            suffix = " DIP" if re.search(r"\bDIP", spec_text.upper()) else ""
            return f"{physical_size.group(1)}x{physical_size.group(2)}{suffix}"

        if re.search(r"\bDIP", spec_text.upper()):
            return "DIP"

        return ""

    @staticmethod
    def _extract_tolerance(full_description):
        match = re.search(r"(\d+(?:\.\d+)?)\s*%", full_description.upper())
        return f"{match.group(1)}%" if match else ""

    @staticmethod
    def _extract_power(spec_text):
        match = re.search(r"\b(\d+(?:/\d+)?(?:\.\d+)?)\s*W\b", spec_text.upper())
        return f"{match.group(1)}W" if match else ""

    @staticmethod
    def _extract_capacitor_technology(spec_text):
        upper_spec = spec_text.upper()
        if "MLCC" in upper_spec or "CAP CER" in upper_spec:
            return "MLCC"
        if "CAP PL" in upper_spec or "POLYMER" in upper_spec:
            return "導電高分子"
        if "TANT" in upper_spec:
            return "鉭質"
        if "ELECT" in upper_spec or "ECAP" in upper_spec:
            return "鋁電解"
        return "電容"

    @staticmethod
    def _extract_resistance(spec_text):
        match = re.search(r"\b(\d+(?:\.\d+)?)\s*([kKmM]?)\s*(?:OHM|Ω)\b", spec_text)
        if match:
            magnitude = float(match.group(1))
            unit = match.group(2)
            multiplier = {"": 1, "K": 1_000, "k": 1_000, "M": 1_000_000, "m": 0.001}[unit]
            return match.group(0).strip(), SixtyBOMAnnotator._format_resistance(magnitude * multiplier), magnitude * multiplier

        match = re.search(r"\b(\d+)([RrKkMm])(\d+)\b", spec_text)
        if match:
            before, unit, after = match.groups()
            magnitude = float(f"{before}.{after}")
            multiplier = {"R": 1, "r": 1, "K": 1_000, "k": 1_000, "M": 1_000_000, "m": 0.001}[unit]
            return match.group(0), SixtyBOMAnnotator._format_resistance(magnitude * multiplier), magnitude * multiplier

        match = re.search(r"\bR(\d+)\b", spec_text.upper())
        if match:
            magnitude = float(f"0.{match.group(1)}")
            return match.group(0), SixtyBOMAnnotator._format_resistance(magnitude), magnitude

        return "", "", None

    @staticmethod
    def _format_resistance(ohms):
        if ohms >= 1_000_000:
            return f"{ohms / 1_000_000:g}MOhm"
        if ohms >= 1_000:
            return f"{ohms / 1_000:g}kOhm"
        if 0 < ohms < 1:
            return f"{ohms * 1_000:g}mOhm"
        return f"{ohms:g}Ohm"

    @staticmethod
    def _capacitance_to_pf(value):
        match = re.fullmatch(r"(\d+(?:\.\d+)?)(pF|nF|uF)", value)
        if not match:
            return None

        magnitude, unit = match.groups()
        return float(magnitude) * {"pF": 1, "nF": 1_000, "uF": 1_000_000}[unit]

    def _link_substitutes(self, records):
        active_primary = None
        groups = defaultdict(list)

        for record in records:
            context = (record["bom_level"], record["seq_num"])
            if record["role"] == "主料":
                record["group_id"] = f"{record['row']}:{record['bom_level']}:{record['seq_num']}"
                active_primary = record
            elif active_primary and context == (active_primary["bom_level"], active_primary["seq_num"]):
                record["group_id"] = active_primary["group_id"]
            else:
                record["group_id"] = f"{record['row']}:{record['bom_level']}:{record['seq_num']}"

            groups[record["group_id"]].append(record)

        for group in groups.values():
            primary = next((record for record in group if record["role"] == "主料"), None)
            if primary is None:
                continue

            substitutes = [record for record in group if record["role"] == "Second Source"]
            primary["substitute_count"] = len(substitutes)
            for substitute in substitutes:
                substitute["substitute_differences"] = []

    def _find_same_specification_candidates(self, records):
        groups = defaultdict(list)
        for record in records:
            if record["role"] == "主料" and self._can_compare(record):
                groups[record["normalize_key"]].append(record)

        for group in groups.values():
            part_numbers = {record["item_number"] for record in group}
            if len(part_numbers) < 2:
                continue

            for record in group:
                record["same_specification_groups"] = len(part_numbers) - 1

    def _assign_normalization_groups(self, records):
        groups = defaultdict(list)
        for record in records:
            record["normalize_group_id"] = ""
            record["normalize_group_members"] = []
            if self._can_compare(record):
                groups[record["normalize_key"]].append(record)

        group_counter = 1
        for normalize_key in sorted(groups):
            group = sorted(groups[normalize_key], key=lambda current: current["row"])
            members = []
            for record in group:
                if record["item_number"] and record["item_number"] not in members:
                    members.append(record["item_number"])

            if len(members) < 2:
                continue

            group_id = f"RC{group_counter:04d}"
            group_counter += 1
            for record in group:
                record["normalize_group_id"] = group_id
                record["normalize_group_members"] = members

    def _find_near_value_candidates(self, records):
        families = defaultdict(list)
        for record in records:
            if record["role"] == "主料" and self._can_compare(record) and record["numeric_value"] is not None:
                families[self._family_key(record)].append(record)

        for family in families.values():
            values = defaultdict(list)
            for record in family:
                values[record["normalized_value"]].append(record)

            sorted_values = sorted(
                (
                    {
                        "normalized_value": normalized_value,
                        "numeric_value": value_records[0]["numeric_value"],
                        "records": value_records,
                    }
                    for normalized_value, value_records in values.items()
                ),
                key=lambda value_group: value_group["numeric_value"],
            )
            for lower, higher in itertools.combinations(sorted_values, 2):
                if lower["numeric_value"] <= 0:
                    continue

                ratio = higher["numeric_value"] / lower["numeric_value"]
                if ratio > self.near_value_ratio:
                    continue

                for lower_record in lower["records"]:
                    lower_record["near_values"].append((higher["normalized_value"], ratio))
                for higher_record in higher["records"]:
                    higher_record["near_values"].append((lower["normalized_value"], ratio))

    @staticmethod
    def _family_key(record):
        if record["component_type"] == "電容":
            return (
                record["component_type"],
                record["technology"],
                record["voltage"],
                record["dielectric"],
                record["package"],
                record["tolerance"],
            )

        return (
            record["component_type"],
            record["package"],
            record["tolerance"],
            record["power"],
        )

    @staticmethod
    def _can_compare(record):
        return bool(
            record["role"] == "主料"
            and record["component_type"]
            and record["normalize_key"]
            and not record["missing_fields"]
        )

    def _set_judgments(self, records):
        for record in records:
            (
                record["judgment"],
                record["suggestion"],
                record["risk_category"],
                record["risk_reason"],
                record["decision_label"],
                record["decision_color"],
            ) = self._judgment_for(record)

    def _judgment_for(self, record):
        if record["role"] == "Second Source":
            return (
                "Second Source（不納入比較）",
                "此料為華碩料庫中主料的既有 Second Source；不納入不同主料之間的歸一化分析。",
                "Second Source",
                "既有替代料，不參與主料之間的規格歸一化比較。",
                "灰色｜Second Source",
                "gray",
            )

        if not record["component_type"]:
            return (
                "不適用",
                "非電阻/電容料件，本輪不進行規格歸一化判定。",
                "Assembly",
                "非本輪分析料型，需人工確認是否納入後續審查。",
                "🟡 建議確認",
                "yellow",
            )

        if record["missing_fields"]:
            missing = "、".join(record["missing_fields"])
            return (
                "規格資訊不足",
                f"缺少{missing}，不可依目前文字自動判定為等效料；請補齊規格後再進行歸一化審查。",
                self._risk_category_for_missing_fields(record["missing_fields"]),
                f"規格欄位缺漏（{missing}），存在誤判歸一化風險。",
                "🔴 禁止統一",
                "red",
            )

        if record["is_critical"]:
            return (
                "關鍵電路保護",
                "描述含高速、CPU、RF 或關鍵電路關鍵字；不建議自動歸一化或替換，須由 RD 設計負責人確認。",
                self._critical_risk_category(record["spec_text"]),
                "偵測到關鍵電路語境，變更可能影響訊號完整性或供電穩定。",
                "🔴 禁止統一",
                "red",
            )

        if record["same_specification_groups"]:
            return (
                "可評估歸一化",
                f"發現{record['same_specification_groups']}個其他主料組使用相同完整規格；可由 RD/採購檢查 AVL、生命週期與佈局後，評估共料或料號整合。{self._second_source_note(record)}",
                "Vendor",
                "同規格群可整併，但仍需確認供應商策略與料況。",
                "🟢 可以統一",
                "green",
            )

        if record["near_values"]:
            return (
                "相近值需確認",
                f"同系列存在相近值；請由 RD 確認電路功能、裕量與容差後再決定是否替代。{self._second_source_note(record)}",
                "Electrical",
                "阻容值接近但非等值，需確認功能與設計裕量。",
                "🟡 建議確認",
                "yellow",
            )

        return (
            "規格已標準化",
            f"未發現跨組相同完整規格或需特別確認的相近值；維持目前料件。{self._second_source_note(record)}",
            "Vendor",
            "目前未發現可直接整併群組，維持現況並持續觀察供應策略。",
            "🟡 建議確認",
            "yellow",
        )

    @staticmethod
    def _risk_category_for_missing_fields(missing_fields):
        missing = set(missing_fields)
        if "封裝/尺寸" in missing:
            return "Package"
        if "功率" in missing or "阻值" in missing or "電容量" in missing or "額定電壓" in missing:
            return "Electrical"
        return "Assembly"

    @staticmethod
    def _critical_risk_category(spec_text):
        upper_spec = spec_text.upper()
        signal_keywords = (
            "PCIE",
            "PCI-E",
            "DDR",
            "USB",
            "TYPE-C",
            "TYPEC",
            "HDMI",
            "DISPLAYPORT",
            "RF",
            "WLAN",
            "WI-FI",
            "WIFI",
            "ANTENNA",
        )
        if any(keyword in upper_spec for keyword in signal_keywords):
            return "Signal Integrity"
        return "Electrical"

    @staticmethod
    def _second_source_note(record):
        if not record["substitute_count"]:
            return ""

        return (
            f" 本主料已有{record['substitute_count']}項 Second Source；"
            "Second Source 不納入主料間歸一化比較。"
        )

    def _is_critical(self, spec_text):
        upper_spec = spec_text.upper()
        return any(re.search(rf"\b{re.escape(keyword)}\b", upper_spec) for keyword in self.CRITICAL_KEYWORDS)

    def _write_annotations(self, worksheet, records):
        header_map = self._header_map(worksheet)
        first_column = max(header_map.values()) + 1
        destination_columns = {}
        highlight_columns = {
            "歸一化群組(Group ID)",
            "群組成員(主料)",
            "相近值提示",
            "歸一化判定",
            "歸一化建議",
            "風險分類",
            "風險原因",
            "判定燈號",
        }
        for offset, header in enumerate(self.ANALYSIS_COLUMNS):
            column = first_column + offset
            destination_columns[header] = column
            cell = worksheet.cell(1, column)
            cell.value = header
            cell.fill = PatternFill("solid", fgColor="1F4E78")
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            worksheet.column_dimensions[cell.column_letter].width = self._column_width(header)

        for record in records:
            values = {
                "元件類型": record["component_type"] or "-",
                "60BOM料件角色": record["role"],
                "原始規格值": record["original_value"] or "-",
                "歸一化值": self._display_normalized_value(
                    record["normalized_value"], record["component_type"]
                ),
                "額定電壓": record["voltage"] or "-",
                "介質": record["dielectric"] or "-",
                "封裝/尺寸": record["package"] or "-",
                "容差": record["tolerance"] or "-",
                "功率": record["power"] or "-",
                "歸一化規格鍵": record["normalize_key"] or "-",
                "歸一化群組(Group ID)": record["normalize_group_id"] or "-",
                "群組成員(主料)": "、".join(record["normalize_group_members"]) if record["normalize_group_members"] else "-",
                "相近值提示": self._near_value_note(record),
                "歸一化判定": record["judgment"],
                "歸一化建議": record["suggestion"],
                "風險分類": record["risk_category"],
                "風險原因": record["risk_reason"],
                "判定燈號": record["decision_label"],
            }
            for header, value in values.items():
                cell = worksheet.cell(record["row"], destination_columns[header])
                cell.value = value
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                if header in highlight_columns:
                    cell.fill = self._decision_fills.get(record["decision_color"], PatternFill())

        worksheet.auto_filter.ref = worksheet.dimensions
        worksheet.freeze_panes = "A2"

    @staticmethod
    def _column_width(header):
        widths = {
            "元件類型": 12,
            "60BOM料件角色": 16,
            "原始規格值": 16,
            "歸一化值": 16,
            "額定電壓": 12,
            "介質": 12,
            "封裝/尺寸": 15,
            "容差": 10,
            "功率": 10,
            "歸一化規格鍵": 42,
            "歸一化群組(Group ID)": 20,
            "群組成員(主料)": 30,
            "相近值提示": 45,
            "歸一化判定": 24,
            "歸一化建議": 68,
            "風險分類": 18,
            "風險原因": 42,
            "判定燈號": 16,
        }
        return widths[header]

    @classmethod
    def _near_value_note(cls, record):
        notes = []

        if record["component_type"] == "電容":
            capacitor_note = cls._capacitor_standard_note(record["normalized_value"])
            if capacitor_note:
                notes.append(capacitor_note)

        if record["component_type"] == "電阻":
            resistor_note = cls._resistor_standard_note(record["normalized_value"])
            if resistor_note:
                notes.append(resistor_note)

        if record["near_values"]:
            nearest = []
            for value, ratio in sorted(record["near_values"], key=lambda item: item[1]):
                compact_value = cls._compact_value(value, record["component_type"])
                note = f"{compact_value} ({ratio:.2f}:1)"
                if note not in nearest:
                    nearest.append(note)
            notes.append(f"BOM內附近值：{'、'.join(nearest[:5])}")

        if not notes:
            return "-"

        notes.append("僅供 RD 判讀，不建議自動替換。")
        return "\n".join(notes)

    @classmethod
    def _capacitor_standard_note(cls, normalized_value):
        capacitance_pf = cls._capacitance_to_pf(normalized_value)
        if capacitance_pf is None or capacitance_pf <= 0:
            return ""

        nearby_values = cls._nearby_values_from_series(
            capacitance_pf,
            cls.E12_VALUES,
            count=3,
            formatter=cls._format_capacitance_pf,
        )
        if not nearby_values:
            return ""

        return (
            f"目前值：{cls._format_capacitance_pf(capacitance_pf)}\n"
            f"標準系列：E12\n"
            f"附近值：{'、'.join(nearby_values)}"
        )

    @classmethod
    def _resistor_standard_note(cls, normalized_value):
        resistance_ohm = cls._resistance_to_ohm(normalized_value)
        if resistance_ohm is None or resistance_ohm <= 0:
            return ""

        primary_series = cls._resistor_series(resistance_ohm)
        primary_nearby = cls._nearby_values_from_series(
            resistance_ohm,
            cls._e_series_values(primary_series),
            count=3,
            formatter=cls._format_resistance_compact,
        )
        fine_nearby = cls._nearby_values_from_series(
            resistance_ohm,
            cls._e_series_values("E96"),
            count=3,
            formatter=cls._format_resistance_compact,
        )
        fine_nearby = [value for value in fine_nearby if value not in primary_nearby]

        lines = [
            f"目前值：{cls._format_resistance_compact(resistance_ohm)}",
            f"標準系列：{primary_series}",
            f"附近值：{'、'.join(primary_nearby) if primary_nearby else '-'}",
        ]
        if fine_nearby:
            lines.append(f"附近精細值(E96)：{'、'.join(fine_nearby[:3])}")
        return "\n".join(lines)

    @classmethod
    def _resistor_series(cls, resistance_ohm):
        mantissa = cls._normalized_mantissa(resistance_ohm)
        if cls._matches_series(mantissa, "E24", significant_digits=2):
            return "E24"
        if cls._matches_series(mantissa, "E96", significant_digits=3):
            return "E96"
        if cls._matches_series(mantissa, "E192", significant_digits=3):
            return "E192"
        return "E24"

    @classmethod
    def _matches_series(cls, mantissa, series_name, significant_digits):
        rounded = cls._round_significant(mantissa, significant_digits)
        if abs(rounded - mantissa) / mantissa > 0.0005:
            return False

        return any(abs(rounded - value) < 1e-9 for value in cls._e_series_values(series_name))

    @classmethod
    def _e_series_values(cls, series_name):
        if series_name in cls._E_SERIES_CACHE:
            return cls._E_SERIES_CACHE[series_name]

        series_parameters = {
            "E24": (24, 2),
            "E96": (96, 3),
            "E192": (192, 3),
        }
        count, digits = series_parameters[series_name]
        values = []
        for index in range(count):
            raw_value = 10 ** (index / count)
            rounded_value = cls._round_significant(raw_value, digits)
            if rounded_value >= 10:
                rounded_value /= 10
            if not values or abs(values[-1] - rounded_value) > 1e-9:
                values.append(rounded_value)

        cls._E_SERIES_CACHE[series_name] = tuple(values)
        return cls._E_SERIES_CACHE[series_name]

    @staticmethod
    def _round_significant(value, digits):
        if value == 0:
            return 0.0
        decimal_places = digits - int(math.floor(math.log10(abs(value)))) - 1
        return round(value, decimal_places)

    @staticmethod
    def _normalized_mantissa(value):
        exponent = math.floor(math.log10(abs(value)))
        return value / (10 ** exponent)

    @classmethod
    def _nearby_values_from_series(cls, numeric_value, series_values, count, formatter):
        if numeric_value <= 0:
            return []

        base_exponent = math.floor(math.log10(numeric_value))
        candidates = []
        for offset in (-1, 0, 1):
            scale = 10 ** (base_exponent + offset)
            for mantissa in series_values:
                candidate = mantissa * scale
                if abs(candidate - numeric_value) < 1e-12:
                    continue
                ratio = abs(candidate - numeric_value) / numeric_value
                candidates.append((ratio, candidate))

        unique = []
        seen = set()
        for _, value in sorted(candidates, key=lambda item: (item[0], item[1])):
            key = round(value, 12)
            if key in seen:
                continue
            seen.add(key)
            unique.append(value)
            if len(unique) >= count:
                break

        return [formatter(value) for value in sorted(unique)]

    @staticmethod
    def _format_capacitance_pf(value_pf):
        if value_pf >= 1_000_000:
            return f"{value_pf / 1_000_000:g}uF"
        if value_pf >= 1_000:
            return f"{value_pf / 1_000:g}nF"
        return f"{value_pf:g}pF"

    @staticmethod
    def _resistance_to_ohm(value):
        match = re.fullmatch(r"(\d+(?:\.\d+)?)(mOhm|Ohm|kOhm|MOhm)", value)
        if not match:
            return None

        magnitude, unit = match.groups()
        multiplier = {
            "mOhm": 0.001,
            "Ohm": 1,
            "kOhm": 1_000,
            "MOhm": 1_000_000,
        }
        return float(magnitude) * multiplier[unit]

    @staticmethod
    def _format_resistance_compact(ohms):
        if ohms >= 1_000_000:
            return f"{ohms / 1_000_000:g}M"
        if ohms >= 1_000:
            return f"{ohms / 1_000:g}K"
        if ohms >= 1:
            return f"{ohms:g}R"
        return f"{ohms * 1_000:g}mR"

    @classmethod
    def _compact_value(cls, normalized_value, component_type):
        if component_type == "電容":
            capacitance_pf = cls._capacitance_to_pf(normalized_value)
            return cls._format_capacitance_pf(capacitance_pf) if capacitance_pf else normalized_value

        if component_type == "電阻":
            resistance_ohm = cls._resistance_to_ohm(normalized_value)
            return cls._format_resistance_compact(resistance_ohm) if resistance_ohm else normalized_value

        return normalized_value

    @staticmethod
    def _display_normalized_value(value, component_type):
        if not value:
            return "-"

        if component_type == "電容":
            match = re.fullmatch(r"(\d+(?:\.\d+)?)(pF|nF|uF)", value)
            if match:
                magnitude, unit = match.groups()
                unit_map = {"pF": "皮法", "nF": "納法", "uF": "微法"}
                return f"{magnitude}{unit_map[unit]}"

        if component_type == "電阻":
            match = re.fullmatch(r"(\d+(?:\.\d+)?)(mOhm|Ohm|kOhm|MOhm)", value)
            if match:
                magnitude, unit = match.groups()
                unit_map = {
                    "mOhm": "毫歐姆",
                    "Ohm": "歐姆",
                    "kOhm": "千歐姆",
                    "MOhm": "兆歐姆",
                }
                return f"{magnitude}{unit_map[unit]}"

        return value

    @staticmethod
    def _summary(records, output_path):
        electronic_records = [record for record in records if record["component_type"]]
        return {
            "output_file": str(output_path),
            "total_rows": len(records),
            "electronic_rows": len(electronic_records),
            "same_specification_candidates": sum(
                1 for record in electronic_records if record["same_specification_groups"]
            ),
            "normalization_groups": len(
                {
                    record["normalize_group_id"]
                    for record in electronic_records
                    if record["normalize_group_id"]
                }
            ),
            "near_value_candidates": sum(
                1 for record in electronic_records if record["near_values"]
            ),
            "substitute_rows": sum(1 for record in records if record["role"] == "Second Source"),
            "critical_rows": sum(1 for record in electronic_records if record["is_critical"]),
        }